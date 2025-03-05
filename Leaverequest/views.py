import openpyxl
import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from UserLogin.models import EmployeeLogin
from .models import LeaveBalances, LeaveType, LeaveCategory, LeaveRequest, LeaveRouting
from Notification.models import Notification 
from django.utils import timezone
from django.core.exceptions import ObjectDoesNotExist
from Calendar.models import Event, EventType
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from datetime import datetime, timedelta
from Profile.models import EmploymentInformation
from rest_framework import generics
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .serializers import LeaveRoutingSerializer
from django.db.models import Count, FloatField
from django.db.models.functions import Cast
from rest_framework.decorators import api_view
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from django.http import HttpResponse
from collections import defaultdict

def is_admin(user):
    return user.is_admin

@login_required(login_url='login')
def leave_view(request):
    current_year = timezone.now().year
    next_year = timezone.now().year + 1
    today = timezone.now().date()

    try:
        vlcurrentyear = LeaveBalances.objects.get(name=request.user, leave_type="VL", date_started__year=current_year)
        slcurrentyear = LeaveBalances.objects.get(name=request.user, leave_type="SL", date_started__year=current_year)
        vlnextyear = LeaveBalances.objects.get(name=request.user, leave_type="VL", date_started__year=next_year)
        slnextyear = LeaveBalances.objects.get(name=request.user, leave_type="SL", date_started__year=next_year)
    except ObjectDoesNotExist:
        vlcurrentyear = None 
        slcurrentyear = None 
        vlnextyear = None
        slnextyear = None

    upcoming_leaves = LeaveRequest.objects.filter(employee=request.user, from_date__gte=today).order_by('from_date')
    next_upcoming_leave = upcoming_leaves.first()
    requests = LeaveRequest.objects.filter(employee=request.user).order_by('-submitted_at')
    for leave_request in requests:
        leave_request_routings = LeaveRouting.objects.filter(leave_request=leave_request.id)  # Filter by leave_request
        if leave_request_routings.exists():
            latest_routing = leave_request_routings.last()  
            leave_request.approver = latest_routing.approver.name
            leave_request.remarks = latest_routing.status  # Assuming you meant status for remarks
        else:
            leave_request.approver = None
            leave_request.remarks = None

    approver = EmploymentInformation.objects.get(name=request.user)
    leavetypes = LeaveType.objects.all()
    context={
        'vlcurrentyear':vlcurrentyear,
        'slcurrentyear':slcurrentyear,
        'vlnextyear':vlnextyear,
        'slnextyear':slnextyear,
        'requests':requests,
        'leavetypes':leavetypes,
        'approver':approver,
        'next_upcoming_leave':next_upcoming_leave
    }
    return render(request, 'Leaverequest/leave-request.html', context)

@login_required(login_url='login')
def get_leave_categories(request):
    leave_type_id = request.GET.get('leave_type_id')
    categories = LeaveCategory.objects.filter(leave_type_id=leave_type_id).values('id', 'category_name')
    category_list = list(categories)
    return JsonResponse(category_list, safe=False)

@login_required(login_url='login')
@require_http_methods(["POST"])
def calculate_leave_days(request):
    try:
        data = json.loads(request.body)
        start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
        end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()
        
        if start_date == end_date:
            if start_date.weekday() < 6:
                try:
                    working_day_type = EventType.objects.get(name="Working day")
                    excluded_event = Event.objects.filter(
                        event_date__date=start_date
                    ).exclude(event_type=working_day_type).exists()
                except EventType.DoesNotExist:
                    excluded_event = Event.objects.filter(
                        event_date__date=start_date
                    ).exists()
                
                working_days = 0 if excluded_event else 1
            else:
                working_days = 0
            
            return JsonResponse({'working_days': working_days})
        
        try:
            working_day_type = EventType.objects.get(name="Working day")
            excluded_events = Event.objects.filter(
                event_date__date__range=(start_date, end_date),
                event_type__isnull=False,
            ).exclude(event_type=working_day_type)
        except EventType.DoesNotExist:
            excluded_events = Event.objects.filter(
                event_date__date__range=(start_date, end_date)
            )
        
        excluded_dates = {event.event_date.date() for event in excluded_events}
        
        working_days = 0
        current_date = start_date
        
        while current_date <= end_date:
            if current_date.weekday() < 6:
                if current_date not in excluded_dates:
                    working_days += 1
            current_date += timedelta(days=1)
        
        return JsonResponse({'working_days': working_days})
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    
@login_required(login_url='login')
def submit_leave(request):
    if request.method == "POST":

        date_from = request.POST.get('leave_from')
        date_to = request.POST.get('leave_to')
        leave_type = request.POST.get('leave_type')
        selectedType = LeaveType.objects.get(id=leave_type)
        
        no_of_days = request.POST.get('no_of_days')
        no_of_hrs = request.POST.get('no_of_hrs')
        category = request.POST.get('leave_category')
        selectedCategory = LeaveCategory.objects.get(id=category)
        reason = request.POST.get('reason')

        current_date = timezone.now().date()
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()

        if selectedType.name in ["Maternity Leave", "Magna Carta", "Sickness Leave"]:
            approver = EmployeeLogin.objects.get(idnumber="Clinic")
            leave_approver = approver
        # elif date_from <= current_date:
        #     approver = EmployeeLogin.objects.get(idnumber="IAD")
        #     leave_approver = approver
        elif date_from == current_date + timedelta(days=1) or date_from == current_date + timedelta(days=2):
            approver = EmploymentInformation.objects.get(name=request.user)
            leave_approver = approver.approver
        else:
            approver = EmployeeLogin.objects.get(idnumber="IAD")
            leave_approver = approver

        leaveForm = LeaveRequest(
            employee=request.user,
            leave_type=selectedType,
            days=no_of_days,
            hrs=no_of_hrs,
            from_date=date_from,
            to_date=date_to,
            category=selectedCategory,
            reason=reason
        )
        leaveForm.save()

        LeaveRouting.objects.create(
            leave_request=leaveForm,
            approver=leave_approver
        )

        Notification.objects.create(
            level="High",
            module="Leave Request",
            notifier=request.user,
            page="leave",
            reciever=leave_approver,
            message=f'has submitted a {selectedType.name}'
        )

        messages.success(request, 'Leave submitted successfully')
        return redirect('leave')

    return redirect('leave')

class LeaveRoutingListAPIView(generics.ListAPIView):
    serializer_class = LeaveRoutingSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        leave_request_id = self.kwargs['leave_request_id']
        return LeaveRouting.objects.filter(leave_request_id=leave_request_id)

@login_required(login_url='login')
def delete_leave(request, pk):
    selectedLeave = get_object_or_404(LeaveRequest, id=pk)
    if request.method == "POST":
        selectedLeave.delete()
        messages.success(request, 'Leave cancelled.')
        return redirect('leave')
    return redirect('leave')

@login_required(login_url='login')
@user_passes_test(is_admin, login_url='Dashboard:Dashboard')
def leave_balances(request):
    if request.user.is_accounting == False:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    balances = LeaveBalances.objects.all().order_by('-name__idnumber')
    context={
        'balances':balances
    }
    return render(request, 'Leaverequest/leave-balances.html', context)

@login_required(login_url='login')
@user_passes_test(is_admin, login_url='Dashboard:Dashboard')
def import_balances(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect('leave-balances')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception as e:
            messages.error(request, f"Error loading file: {e}")
            return redirect('leave-balances')

        LeaveBalances.objects.all().delete()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                idnumber = row[0]
                leave_type = row[1]
                date_started = row[2]
                date_ended = row[3]
                balances = row[4]

                employee = EmployeeLogin.objects.get(idnumber=idnumber)

                LeaveBalances.objects.create(
                    name=employee,
                    leave_type=leave_type,
                    date_started=date_started,
                    date_ended=date_ended,
                    balances=balances,
                )

            except EmployeeLogin.DoesNotExist:
                messages.error(request, f"Employee with ID {idnumber} does not exist.")
            except Exception as e:
                messages.error(request, f"Error processing employee {idnumber}: {e}")

        messages.success(request, "Balances uploaded successfully.")
        return redirect('leave-balances')

    return redirect('leave-balances')

@login_required(login_url='login')
def admin_approval(request):
    requests = LeaveRouting.objects.filter(approver=request.user, status="Waiting for Approval").order_by('-request_at')

    context={
        'requests':requests
    }
    return render(request, 'Leaverequest/admin-approval.html', context)

@login_required(login_url='login')
def admin_leave_approval(request, pk):
    selectedLeave = get_object_or_404(LeaveRouting, id=pk)
    leavestatus = get_object_or_404(LeaveRequest, id=selectedLeave.leave_request.id)
    approver_level = EmploymentInformation.objects.get(name=request.user)
    
    if request.method == 'POST':
        remarks = request.POST.get('approver_remarks')
        action = request.POST.get('action') 

        if request.user.idnumber == "Clinic":
            approver = EmployeeLogin.objects.get(idnumber="IAD")
            leave_approver = approver
        elif request.user.idnumber == "IAD":
            approver = EmploymentInformation.objects.get(name=selectedLeave.leave_request.employee)
            leave_approver = approver.approver
        elif request.user.idnumber == "HRD":
            leave_approver = None
            
        elif approver_level.position.approver_level == 2:   
            approver = EmployeeLogin.objects.get(idnumber='HRD')
            leave_approver = approver
        else:
            approver = EmploymentInformation.objects.get(name=request.user)
            leave_approver = approver.approver

        if remarks:
            selectedLeave.remarks = remarks

        if action == 'approve':
            selectedLeave.status = "Approved"
            message_text = 'Leave approved.'
        elif action == 'disapprove':
            selectedLeave.status = "Disapproved"
            message_text = 'Leave disapproved.'

        selectedLeave.approved_at = timezone.now()
        selectedLeave.save()

        if request.user.idnumber != "HRD":
            LeaveRouting.objects.create(
                leave_request=selectedLeave.leave_request,
                approver=leave_approver
            )

        if request.user.idnumber == "HRD":
            
            if action == 'approve':
                leavestatus.status = "Approved"
            elif action == 'disapprove':
                leavestatus.status = "Disapproved"
            
            leavestatus.save()
                
            Notification.objects.create(
                level="High",
                module="Leave Request",
                notifier=request.user,
                page="leave",
                reciever=selectedLeave.leave_request.employee,
                message=f'has {selectedLeave.status.lower()} your {selectedLeave.leave_request.leave_type.name.lower()}'
            )
        else:
            Notification.objects.create(
                level="High",
                module="Leave Request",
                notifier=request.user,
                page="admin-approval",
                reciever=leave_approver,
                message=f'has {selectedLeave.status.lower()} {selectedLeave.leave_request.employee.name.lower()} {selectedLeave.leave_request.leave_type.name.lower()}'
            )

        messages.success(request, message_text)
        return redirect('admin-approval')

    return redirect('admin-approval')


@login_required(login_url='login')
def leave_history(request):
    requests = LeaveRouting.objects.filter(approver=request.user).order_by('-request_at')
    context={
        'requests':requests
    }
    return render(request, 'Leaverequest/leave-history.html', context)

@api_view(['GET'])
def department_percentages_leave(request):
    total_leave = LeaveRouting.objects.filter(approver=request.user).count()

    if total_leave == 0:
        return Response({
            'labels': [],
            'percentages': [],
            'total_leave': total_leave
        })

    department_percentages = LeaveRouting.objects.filter(
        approver=request.user,
        leave_request__employee__employment_info__department__abreviation__isnull=False
    ).values(
        'leave_request__employee__employment_info__department__abreviation'
    ).annotate(
        count=Count('id'),
        percentage=Cast(Count('id') * 100.0 / total_leave, FloatField())
    ).order_by('leave_request__employee__employment_info__department__abreviation')

    labels = [dept['leave_request__employee__employment_info__department__abreviation'] for dept in department_percentages]
    percentages = [round(dept['percentage'], 1) for dept in department_percentages]

    return Response({
        'labels': labels,
        'percentages': percentages,
        'total_leave': total_leave
    })

@login_required(login_url='login')
@user_passes_test(is_admin, login_url='Dashboard:Dashboard')
def admin_leavelist(request):
    requests = LeaveRequest.objects.all().order_by('-submitted_at')
    leave_types = LeaveType.objects.all()
    categories = LeaveCategory.objects.all()
    context={
        'requests':requests,
        'leave_types':leave_types,
        'categories':categories
    }
    return render(request, 'Leaverequest/admin-leavelist.html', context)


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='Dashboard:Dashboard')
def export_leave_excel(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if request.GET and (not from_date or not to_date):
        messages.error(request, "Please select both 'From' and 'To' dates to filter.")
        return redirect('admin-leave')
    
    requests = LeaveRequest.objects.all()

    if from_date and to_date:
        try:
            from_date_parsed = datetime.strptime(from_date, '%d/%m/%Y')
            to_date_parsed = datetime.strptime(to_date, '%d/%m/%Y')
            
            to_date_parsed = to_date_parsed.replace(hour=23, minute=59, second=59)
            requests = requests.filter(
                submitted_at__range=(from_date_parsed, to_date_parsed)
            )
        except ValueError:
            messages.error(request, "Invalid date format. Please use DD/MM/YYYY format.")
            return redirect('admin-leave')

    total_requests_count = requests.count()

    leave_type_counts = defaultdict(int)
    leave_category_counts = defaultdict(int)

    for request in requests:
        leave_type_counts[request.leave_type.name] += 1
        leave_category_counts[request.category.category_name] += 1

    wb = Workbook()
    # Summary Sheet
    ws = wb.active
    ws.title = "Summary"
    ws['A1'] = "Ryonan Electric Philippines Corporation"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = f"Leave Summary Report"
    ws['A4'] = "Leave Types"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = "Percentage"
    ws['B4'].font = Font(bold=True)

    # Write leave type percentages
    for i, (leave_type, count) in enumerate(leave_type_counts.items(), start=5):
        percentage = (count / total_requests_count) * 100 if total_requests_count > 0 else 0
        ws[f'A{i}'] = leave_type
        ws[f'B{i}'] = f"{round(percentage, 2)}%"
    
    # Leave Category Percentages
    ws.append([])  # Add a blank line for separation
    ws.append(["Leave Categories", "Percentage"])
    ws['A' + str(ws.max_row)].font = Font(bold=True)
    ws['B' + str(ws.max_row)].font = Font(bold=True)

    for i, (category, count) in enumerate(leave_category_counts.items(), start=ws.max_row + 1):
        percentage = (count / total_requests_count) * 100 if total_requests_count > 0 else 0
        ws[f'A{i}'] = category
        ws[f'B{i}'] = f"{round(percentage, 2)}%"

    # Format cells
    for row in ws['A4:B' + str(ws.max_row)]:
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Participants Sheet (Your existing code)
    ws = wb.create_sheet("Leave Summary")
    ws['A1'] = "Ryonan Electric Philippines Corporation"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = "Leave Summary Report"
    ws['A5'] = "ID number"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = "Name"
    ws['B5'].font = Font(bold=True)
    ws['C5'] = "Department"
    ws['C5'].font = Font(bold=True)
    ws['D5'] = "Line"
    ws['D5'].font = Font(bold=True)
    ws['E5'] = "From"
    ws['E5'].font = Font(bold=True)
    ws['F5'] = "To"
    ws['F5'].font = Font(bold=True)
    ws['G5'] = "No of Days"
    ws['G5'].font = Font(bold=True)
    ws['H5'] = "No of Hrs"
    ws['H5'].font = Font(bold=True)
    ws['I5'] = "Leave Type"
    ws['I5'].font = Font(bold=True)
    ws['J5'] = "Category"
    ws['J5'].font = Font(bold=True)
    ws['K5'] = "Reason"
    ws['K5'].font = Font(bold=True)

    for i, request in enumerate(requests, start=6):
        request_department = EmploymentInformation.objects.filter(name=request.employee).first()
        request_line = EmploymentInformation.objects.filter(name=request.employee).first()
        ws[f'A{i}'] = request.employee.idnumber
        ws[f'B{i}'] = request.employee.name
        ws[f'C{i}'] = request_department.department.abreviation if request_department else None
        ws[f'D{i}'] = request_line.line.line if request_line else None
        ws[f'E{i}'] = request.from_date
        ws[f'F{i}'] = request.to_date
        ws[f'G{i}'] = request.days
        ws[f'H{i}'] = request.hrs
        ws[f'I{i}'] = request.leave_type.name
        ws[f'J{i}'] = request.category.category_name
        ws[f'K{i}'] = request.reason

    for row in ws['A5:K' + str(5 + len(requests))]:
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=leave_Report_Summary.xlsx'
    wb.save(response)
    
    return response


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='Dashboard:Dashboard')
def export_pending_leave(request):

    wb = Workbook() 
    ws = wb.active
    ws.title = "Pending Leave"
    ws['A1'] = "Ryonan Electric Philippines Corporation"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = "Pending Leave Summary"
    ws['A4'] = "ID number"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = "Name"
    ws['B4'].font = Font(bold=True)
    ws['C4'] = "Department"
    ws['C4'].font = Font(bold=True)
    ws['D4'] = "Line"
    ws['D4'].font = Font(bold=True)
    ws['E4'] = "Leave Type"
    ws['E4'].font = Font(bold=True)
    ws['F4'] = "Category"
    ws['F4'].font = Font(bold=True)
    ws['G4'] = "Approver"
    ws['G4'].font = Font(bold=True)
    
    pendings = LeaveRouting.objects.filter(status="Waiting for Approval").exclude(approver__is_admin=True)
    for pending in pendings:
        try:
            pending_department = EmploymentInformation.objects.filter(name=pending.leave_request.employee).first()
            if pending_department:
                pending.department = pending_department.department.abreviation
            else:
                pending.department = None
            pending_line = EmploymentInformation.objects.filter(name=pending.leave_request.employee).first()
            if pending_line:
                pending.line = pending_line.line.line
            else:
                pending.line = None
        except EmploymentInformation.DoesNotExist:
            pending.department = None
            pending.line = None

    for i, pending in enumerate(pendings, start=5):
        ws[f'A{i}'] = pending.leave_request.employee.idnumber
        ws[f'B{i}'] = pending.leave_request.employee.name
        ws[f'C{i}'] = pending.department
        ws[f'D{i}'] = pending.line
        ws[f'E{i}'] = pending.leave_request.leave_type.name
        ws[f'F{i}'] = pending.leave_request.category.category_name
        ws[f'G{i}'] = pending.approver.name
    
    for row in ws['A4:G' + str(5 + len(pendings))]:
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
     
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Pending_leave_summary.xlsx'
    wb.save(response)
    
    return response

@api_view(['GET'])
def leave_request_percentage_view(request):
    current_month = timezone.now().month
    current_year = timezone.now().year

    monthly_leaves = LeaveRequest.objects.filter(
        submitted_at__year=current_year,
        submitted_at__month=current_month
    )
    
    total_leaves = monthly_leaves.count()
    leave_type_percentages = {}
    if total_leaves > 0:
        leave_counts = monthly_leaves.values('leave_type__name').annotate(count=Count('id'))
        for leave in leave_counts:
            leave_type_name = leave['leave_type__name']
            abbreviation = ''.join(word[0] for word in leave_type_name.split())
            percentage = (leave['count'] / total_leaves) * 100
            leave_type_percentages[abbreviation] = round(percentage, 2)
    
    leave_types = ['VL', 'VL w/o Pay', 'SL', 'PL', 'ML', 'MC', 'BR', 'SPT'] 
    data = [leave_type_percentages.get(lt, 0) for lt in leave_types]

    return Response({
        'labels': leave_types,
        'data': data
    })