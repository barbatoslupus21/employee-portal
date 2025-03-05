from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import UserFeedback
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

@login_required(login_url='login')
def feedback_view(request):
    return render(request, 'Feedback/feedback.html')

@login_required(login_url='login')
def submit_feedback(request):
    if request.method == "POST":
        feedback = request.POST.get('feedback')
        identity = request.POST.get('identity')

        if identity:
            employee = "Anonymous"
        else:
            employee = request.user.name

        FeedbackForm = UserFeedback(
            feedback=feedback,
            employee=employee
        )
        FeedbackForm.save()
        messages.success(request, 'Feedback successfully submitted')
        return redirect('feedback')

    return render(request, 'Feedback/feedback.html')

@login_required(login_url='login')
def admin_feedback(request):
    if not request.user.is_hrmanager:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    feedbacks = UserFeedback.objects.all().order_by("-submitted_at")
    return render(request, 'Feedback/admin-feedback.html', {'feedbacks':feedbacks})

@login_required(login_url='login')
def export_feedback(request):
    wb = Workbook()
    
    # Create a Remarks Sheet
    ws = wb.active
    ws.title = "Remarks"
    
    ws['A1'] = "Ryonan Electric Philippines"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = "Feedbacks and Suggestions"
    ws['A4'] = "Date Submitted"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = "Employee Name"
    ws['B4'].font = Font(bold=True)
    ws['C4'] = "Feedback and Suggestions"
    ws['C4'].font = Font(bold=True)
    
    feedbacks = UserFeedback.objects.all().order_by("-submitted_at")
    
    # Populate feedbacks into the worksheet
    for i, feedback in enumerate(feedbacks, start=5):
        ws[f'A{i}'] = feedback.submitted_at.strftime("%m/%d/%Y")  # Format the date
        ws[f'B{i}'] = feedback.employee
        ws[f'C{i}'] = feedback.feedback
    
    # Apply borders to header and data rows
    for row in ws['A4:C' + str(4 + len(feedbacks))]:
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
       
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Prepare the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Feedback_Report.xlsx'
    wb.save(response)
    
    return response