from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from UserLogin.models import EmployeeLogin
from .models import SurveyForm, Quarter, UserResponse
from datetime import date
from django.db.models.signals import post_save
from django.dispatch import receiver
from django.db.models import Avg, F, Count, Q
from django.http import JsonResponse
from rest_framework.decorators import api_view
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView
from rest_framework.response import Response
from Profile.models import EmploymentInformation
from Notification.models import Notification

@receiver(post_save, sender=Quarter)
def create_responses_for_users(sender, instance, created, **kwargs):
    if created:
        employees = EmployeeLogin.objects.filter(is_admin=False, is_active=True)
        for employee in employees:
            UserResponse.objects.create(employee=employee, quarter=instance)
            Notification.objects.create(
                    level="Medium",
                    module='Survey',
                    notifier=instance.created_by,
                    page="survey",
                    reciever=employee,
                    message='has posted the quarterly survey for you to fill out.'
                )

@login_required(login_url='login')
def survey_view(request):
    today = date.today()
    start_date = date(today.year, 5, 1)
    end_date = date(today.year + 1, 4, 30)

    quarters = UserResponse.objects.filter(posted_at__gte=start_date, posted_at__lte=end_date, employee=request.user)

    context = {
        'quarters': quarters,
    }
    return render(request, 'Survey/survey.html', context)

@login_required(login_url='login')
def admin_survey(request):
    if not request.user.is_iad:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    today = date.today()
    start_date = date(today.year, 5, 1)
    end_date = date(today.year + 1, 4, 30)

    quarters = Quarter.objects.filter(posted_at__gte=start_date, posted_at__lte=end_date)

    context = {
        'quarters': quarters
    }
    return render(request, 'Survey/admin-survey.html', context)

@login_required(login_url='login')
def submit_quarter(request):
    today = date.today()
    start_date = date(today.year, 5, 1)
    end_date = date(today.year + 1, 4, 30)

    quarters = Quarter.objects.filter(posted_at__gte=start_date, posted_at__lte=end_date)
    if request.method == 'POST':
        quarter = request.POST.get('quarter')
        period = request.POST.get('period')

        QuarterForm = Quarter(
            quarter=quarter,
            period=period,
            created_by=request.user
        )

        QuarterForm.save()
        messages.success(request, 'Survey posted successfully.')
        return redirect('admin-survey')
    
    context = {
        'quarters': quarters
    }
    return render(request, 'Survey/admin-survey.html', context)

@login_required(login_url='login')
def edit_survey(request, pk):
    selectedQuarter = get_object_or_404(Quarter, id=pk)

    if request.method == 'POST':

        quarter = request.POST.get('quarter')
        period = request.POST.get('period')
        selectedQuarter.quarter=quarter
        selectedQuarter.period=period

        selectedQuarter.save()
        messages.success(request, 'Survey changes saved.')
        return redirect('admin-survey')
    
    return render(request, 'Survey/admin-survey.html', {'selectedQuarter':selectedQuarter})

@login_required(login_url='login')
def delete_survey(request, pk):
    deletedSurvey = get_object_or_404(Quarter, id=pk)

    if request.method == "POST":
        deletedSurvey.delete()
        messages.success(request, 'Survey deleted.')
        return redirect('admin-survey')
    else:
        messages.error(request, 'Invalid request method.')
    return redirect('admin-survey')

@login_required(login_url='login')
def take_survey(request, pk):
    selectedSurvey = get_object_or_404(Quarter, id=pk)
    return render(request, 'Survey/take-survey.html', {'selectedSurvey':selectedSurvey})

@login_required(login_url='login')
def survey_response(request, pk):
    response = get_object_or_404(UserResponse, quarter=pk, employee=request.user)

    if request.method == 'POST':
        skills = request.POST.get('skills')
        knowledge_of_job = request.POST.get('knowledge')
        orientation_of_job = request.POST.get('orientation')
        quality_of_supervision = request.POST.get('supervision')
        training_and_development = request.POST.get('training')
        job_description = request.POST.get('job_description')
        opportunity_for_advancement = request.POST.get('advancement')
        workload = request.POST.get('workload')

        policy = request.POST.get('policy')

        salary = request.POST.get('salary')
        salary_increase = request.POST.get('salary_increase')

        clinic = request.POST.get('Clinic')
        kiddie_garden = request.POST.get('Kiddie')
        shuttle_service = request.POST.get('Shuttle')
        locker_room = request.POST.get('Locker')
        working_condition = request.POST.get('Working')
        workplace = request.POST.get('Workplace')
        comfort_room = request.POST.get('Comfort')
        canteen = request.POST.get('Canteen')

        summer_outing = request.POST.get('Summer')
        birthday_celebration = request.POST.get('Birthday')
        christmas_party = request.POST.get('Christmas')
        team_building = request.POST.get('team_building')

        remark = request.POST.get('remarks')
        suggestions = request.POST.get('suggestions')


        survey_form = SurveyForm(
            employee = request.user,
            survey_to = response,
            skills=skills,
            knowledge_of_job=knowledge_of_job,
            orientation_of_job=orientation_of_job,
            quality_of_supervision=quality_of_supervision,
            training_and_development=training_and_development,
            job_description=job_description,
            opportunity_for_advancement=opportunity_for_advancement,
            workload=workload,
            policy=policy,
            salary=salary,
            salary_increase=salary_increase,
            clinic=clinic,
            kiddie_garden=kiddie_garden,
            shuttle_service=shuttle_service,
            locker_room=locker_room,
            working_condition=working_condition,
            workplace=workplace,
            comfort_room=comfort_room,
            canteen=canteen,
            summer_outing=summer_outing,
            birthday_celebration=birthday_celebration,
            christmas_party=christmas_party,
            team_building=team_building,
            remarks=remark,
            suggestions=suggestions
        )

        response.action = True
        survey_form.save()
        response.save()
        messages.success(request, 'Your responses has successfully submitted.')
        return redirect('survey')

    return render(request, 'Survey/take-survey.html')

@login_required(login_url='login')
def admin_survey_view(request, pk):

    selectedQuarter = UserResponse.objects.filter(quarter__id=pk)
    quarter = get_object_or_404(Quarter, id=pk)

    respondents = UserResponse.objects.filter(quarter=pk)
    for respondent in respondents:
        try:
            respondent_department = EmploymentInformation.objects.filter(name=respondent.employee).first()
            if respondent_department:
                respondent.department = respondent_department.department.abreviation
            else:
                respondent.department = None
            respondent_line = EmploymentInformation.objects.filter(name=respondent.employee).first()
            if respondent_line:
                respondent.line = respondent_department.line.line
            else:
                respondent.line = None
        except EmploymentInformation.DoesNotExist:
            respondent.department = None
            respondent.line = None

    survey_responses = SurveyForm.objects.filter(survey_to__quarter__id=pk)

    if not survey_responses.exists():
        messages.error(request, 'There are no survey responses available for this quarter.')
        return redirect('admin-survey')

    employeeCount = EmployeeLogin.objects.exclude(is_admin=True).count()
    formatted_employeeCount = f"{employeeCount:04}"
    
    respondent = UserResponse.objects.filter(quarter__id=pk).exclude(action=False).count()
    formatted_respondent = f"{respondent:04}"
    
    survey_fields = [f.name for f in SurveyForm._meta.get_fields() if f.name not in ['id', 'employee', 'survey_to', 'remarks', 'suggestions', 'submitted_at']]
    
    average_scores = SurveyForm.objects.filter(survey_to__quarter__id=pk).aggregate(
        **{f'{field}_avg': Avg(field) for field in survey_fields}
    )
    
    total_average = sum(average_scores.values()) / len(average_scores)
    average_percentage = (total_average / 5) * 100
    
    if employeeCount > 0:
        response_percentage = (respondent / employeeCount) * 100
    else:
        response_percentage = 0
    
    context = {
        'selectedQuarter': selectedQuarter,
        'employeeCount': formatted_employeeCount,
        'respondent': formatted_respondent,
        'response_percentage': round(response_percentage, 2),
        'average_percentage': round(average_percentage, 2),
        'respondents':respondents,
        'survey_responses':survey_responses,
        'quarter':quarter
    }
    
    return render(request, 'Survey/admin-survey-view.html', context)

@api_view(['GET'])
def job_satisfaction(request, pk):
    fields = [
        'skills', 'knowledge_of_job', 'orientation_of_job', 'quality_of_supervision',
        'training_and_development', 'job_description', 'opportunity_for_advancement', 'workload'
    ]
    
    averages = SurveyForm.objects.filter(survey_to__quarter__id=pk).aggregate(
        **{f'{field}_avg': Avg(field) for field in fields}
    )

    percentages = {field: round((averages[f'{field}_avg'] / 5) * 100, 2) if averages[f'{field}_avg'] is not None else 0 
                   for field in fields}
    
    return JsonResponse(percentages)

@api_view(['GET'])
def policy_salary(request, pk):
    fields = [
        'policy', 'salary', 'salary_increase'
    ]
    
    averages = SurveyForm.objects.filter(survey_to__quarter__id=pk).aggregate(
        **{f'{field}_avg': Avg(field) for field in fields}
    )
    
    percentages = {field: round((averages[f'{field}_avg'] / 5) * 100, 2) if averages[f'{field}_avg'] is not None else 0 
                   for field in fields}
    
    return JsonResponse(percentages)

@api_view(['GET'])
def facilities(request, pk):
    fields = [
        'clinic', 'kiddie_garden', 'shuttle_service','locker_room', 'working_condition', 'workplace','comfort_room', 'canteen',
    ]
    
    averages = SurveyForm.objects.filter(survey_to__quarter__id=pk).aggregate(
        **{f'{field}_avg': Avg(field) for field in fields}
    )
    
    percentages = {field: round((averages[f'{field}_avg'] / 5) * 100, 2) if averages[f'{field}_avg'] is not None else 0 
                   for field in fields}
    
    return JsonResponse(percentages)

@api_view(['GET'])
def relation_program(request, pk):
    fields = [
        'summer_outing', 'birthday_celebration', 'christmas_party','team_building'
    ]
    
    averages = SurveyForm.objects.filter(survey_to__quarter__id=pk).aggregate(
        **{f'{field}_avg': Avg(field) for field in fields}
    )
    
    percentages = {field: round((averages[f'{field}_avg'] / 5) * 100, 2) if averages[f'{field}_avg'] is not None else 0 
                   for field in fields}
    
    return JsonResponse(percentages)

@login_required(login_url='login')
def close_survey(request, pk):
    selectedQuarter = get_object_or_404(Quarter, id=pk)

    if request.method == 'POST':
        selectedQuarter.state = "CLOSED"
        selectedQuarter.save()
        messages.success(request, 'Survey closed.')
        return redirect('admin-survey')
    return redirect('admin-survey')


@login_required(login_url='login')
def export_survey_excel(request, pk):
    selectedQuarter = UserResponse.objects.filter(quarter__id=pk)
    quarter = get_object_or_404(Quarter, id=pk)
    wb = Workbook()
    
    # Summary Sheet
    ws = wb.active
    ws.title = "Summary"
    ws['A1'] = "Ryonan Electric Philippines"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = f"Job Satisfaction Survey for {quarter.quarter}"
    ws['A4'] = "Fields"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = "Average Percentage"
    ws['B4'].font = Font(bold=True)
    
    survey_fields = [f.name for f in SurveyForm._meta.get_fields() if f.name not in ['id', 'employee', 'survey_to', 'remarks', 'suggestions', 'submitted_at']]
    average_scores = SurveyForm.objects.filter(survey_to__quarter__id=pk).aggregate(
        **{f'{field}_avg': Avg(field) for field in survey_fields}
    )
    
    for i, field in enumerate(survey_fields, start=5):
        ws[f'A{i}'] = field.replace('_', ' ').title()
        ws[f'B{i}'] = f"{round((average_scores[f'{field}_avg'] / 5) * 100, 2)}%"
    
    for row in ws['A4:B' + str(4 + len(survey_fields))]:
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Remarks Sheet
    ws = wb.create_sheet("Remarks")
    ws['A1'] = "Ryonan Electric Philippines"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = f"Job Satisfaction Survey for {quarter.quarter}"
    ws['A3'] = "Remarks"
    ws['A5'] = "Name"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = "Remarks"
    ws['B5'].font = Font(bold=True)
    
    survey_responses = SurveyForm.objects.filter(survey_to__quarter__id=pk)
    for i, response in enumerate(survey_responses, start=6):
        ws[f'A{i}'] = response.employee.name
        ws[f'B{i}'] = response.remarks
    
    for row in ws['A5:B' + str(5 + len(survey_responses))]:
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Suggestions Sheet
    ws = wb.create_sheet("Suggestions")
    ws['A1'] = "Ryonan Electric Philippines"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = f"Job Satisfaction Survey for {quarter.quarter}"
    ws['A3'] = "Suggestions"
    ws['A5'] = "Name"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = "Suggestions"
    ws['B5'].font = Font(bold=True)
    
    for i, response in enumerate(survey_responses, start=6):
        ws[f'A{i}'] = response.employee.name
        ws[f'B{i}'] = response.suggestions
    
    for row in ws['A5:B' + str(5 + len(survey_responses))]:
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Respondents Sheet
    ws = wb.create_sheet("Respondents")
    ws['A1'] = "Ryonan Electric Philippines"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = f"Job Satisfaction Survey for {quarter.quarter}"
    ws['A3'] = "Respondents"
    ws['A5'] = "ID number"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = "Name"
    ws['B5'].font = Font(bold=True)
    ws['C5'] = "Department"
    ws['C5'].font = Font(bold=True)
    ws['D5'] = "Line"
    ws['D5'].font = Font(bold=True)
    ws['E5'] = "Email"
    ws['E5'].font = Font(bold=True)
    
    respondents = UserResponse.objects.filter(quarter__id=pk, action=True)
    for respondent in respondents:
        try:
            respondent_department = EmploymentInformation.objects.filter(name=respondent.employee).first()
            if respondent_department:
                respondent.department = respondent_department.department.abreviation
            else:
                respondent.department = None
            respondent_line = EmploymentInformation.objects.filter(name=respondent.employee).first()
            if respondent_line:
                respondent.line = respondent_department.line.line
            else:
                respondent.line = None
        except EmploymentInformation.DoesNotExist:
            respondent.department = None
            respondent.line = None

    for i, respondent in enumerate(respondents, start=6):
        ws[f'A{i}'] = respondent.employee.idnumber
        ws[f'B{i}'] = respondent.employee.name
        ws[f'C{i}'] = respondent.department
        ws[f'D{i}'] = respondent.line
        ws[f'E{i}'] = respondent.employee.email
    
    for row in ws['A5:E' + str(5 + len(respondents))]:
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Survey_Report_{quarter.quarter}.xlsx'
    wb.save(response)
    
    return response

class SurveyRatingPercentagesView(APIView):
    def get(self, request):
        latest_quarter = Quarter.objects.latest('posted_at')
        user_responses = UserResponse.objects.filter(quarter=latest_quarter)
        survey_forms = SurveyForm.objects.filter(survey_to__in=user_responses)
        total_ratings = survey_forms.count() * 23 
        rating_fields = [
            'skills', 'knowledge_of_job', 'orientation_of_job', 'quality_of_supervision',
            'training_and_development', 'job_description', 'opportunity_for_advancement',
            'workload', 'policy', 'salary', 'salary_increase', 'clinic', 'kiddie_garden',
            'shuttle_service', 'locker_room', 'working_condition', 'workplace',
            'comfort_room', 'canteen', 'summer_outing', 'birthday_celebration',
            'christmas_party', 'team_building'
        ]

        rating_counts = {i: 0 for i in range(1, 6)}
        for field in rating_fields:
            for i in range(1, 6):
                rating_counts[i] += survey_forms.filter(**{f"{field}": i}).count()
        
        percentages = {
            'Need Improvement': round(rating_counts[1] / total_ratings * 100, 2),
            'Poor': round(rating_counts[2] / total_ratings * 100, 2),
            'Good': round(rating_counts[3] / total_ratings * 100, 2),
            'Very Good': round(rating_counts[4] / total_ratings * 100, 2),
            'Excellent': round(rating_counts[5] / total_ratings * 100, 2),
        }
        
        overall_average = survey_forms.aggregate(
            avg=Avg(rating_fields[0])
        )['avg']
        for field in rating_fields[1:]:
            overall_average += survey_forms.aggregate(avg=Avg(field))['avg']
        overall_average /= len(rating_fields)
        
        overall_average_percentage = round(overall_average / 5 * 100, 2)
        
        return Response({
            'percentages': percentages,
            'overall_average': overall_average_percentage
        })