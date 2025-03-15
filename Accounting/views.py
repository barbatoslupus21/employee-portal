import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from UserLogin.models import EmployeeLogin
from Profile.models import EmploymentInformation
from .models import FinanceType, Loan, Savings, Medicine, PerfectAttendance, EmployeePayslip
from Profile.models import PersonalInformation
from Notification.models import Notification
from django.core.exceptions import ObjectDoesNotExist
import os
from datetime import datetime
from django.http import JsonResponse, HttpResponseRedirect, HttpResponseNotAllowed
from django.core.mail import EmailMessage
from django.conf import settings
from smtplib import SMTPException
from PyPDF2 import PdfReader, PdfWriter
import tempfile
from django.utils import timezone

def is_admin(user):
    return user.is_admin

@login_required(login_url='login')
@user_passes_test(is_admin, login_url='Dashboard:Dashboard')
def admin_finance(request):
    if request.user.is_accounting == False:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    employees = EmployeeLogin.objects.filter(is_admin=False, is_active=True).order_by('idnumber')
    for employee in employees:
        try:
            employee_department = EmploymentInformation.objects.filter(name=employee).first()
            if employee_department:
                employee.department = employee_department.department.abreviation
            else:
                employee.department = None
            employee_line = EmploymentInformation.objects.filter(name=employee).first()
            if employee_line:
                employee.line = employee_line.line.line
            else:
                employee.line = None
        except EmploymentInformation.DoesNotExist:
            employee.department = None
            employee.line = None
    return render(request, 'Accounting/admin-finance.html', {'employees':employees})

@login_required(login_url='login')
def admin_loan_allowances(request, pk):
    employee = get_object_or_404(EmployeeLogin, id=pk)
    emploans = Loan.objects.filter(employee=pk)
    for emploan in emploans:
        try:
            total_amount = float(emploan.loan_amount) if emploan.loan_amount else 0
            total_payment = float(emploan.payment) if emploan.payment else 0
            emploan.balance = total_amount - total_payment

            emploan.balance_format = "{:,.2f}".format(emploan.balance)
            emploan.loan_amount_format = "{:,.2f}".format(total_amount)
            emploan.payment_format = "{:,.2f}".format(total_payment)
        except ValueError:
            emploan.balance = 0
            emploan.balance_format = "0.00"
            emploan.loan_amount_format = "0.00"
            emploan.payment_format = "0.00"

        if emploan.loan_amount and total_amount > 0:
            emploan.percentage = (emploan.balance / total_amount) * 100
        else:
            emploan.percentage = 0


    savings = Savings.objects.filter(employee=pk)
    for saving in savings:
        try:
            desired_savings = float(saving.desired_savings) if saving.desired_savings else 0
            actual_savings = float(saving.savings) if saving.savings else 0
            saving.balances = desired_savings - actual_savings
            saving.formatted_savings = "{:,.2f}".format(saving.balances)
            saving.desired_savings_format = "{:,.2f}".format(desired_savings)
            saving.actual_savings_format = "{:,.2f}".format(actual_savings)
        except ValueError:
            saving.balances = 0
            saving.formatted_savings = "0.00" 
            saving.desired_savings_format = "0.00" 
            saving.actual_savings_format = "0.00" 

    medicines = Medicine.objects.filter(employee=pk)
    attendances = PerfectAttendance.objects.filter(employee=pk)

    context={
        'emploans':emploans,  
        'savings':savings,   
        'medicines':medicines,
        'attendances':attendances,
        'employee':employee
    }

    return render(request, 'Accounting/admin-loans.html', context)

@login_required(login_url='login')
@user_passes_test(is_admin, login_url='Dashboard:Dashboard')
def import_perfect_attendance(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect('admin-finance')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception as e:
            messages.error(request, f"Error loading file: {e}")
            return redirect('admin-finance')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                idnumber = row[0]
                finance_type_name = row[1]
                title = row[2]
                started_at = row[3]
                ended_at = row[4]
                amount = row[5]
                date_credited = row[6]

                employee = EmployeeLogin.objects.get(idnumber=idnumber)
                finance_type = FinanceType.objects.get(name=finance_type_name) if finance_type_name else None

                PerfectAttendance.objects.create(
                    employee=employee,
                    type=finance_type,
                    title=title,
                    started_at=started_at,
                    ended_at=ended_at,
                    amount=amount,
                    is_credited=True,
                    date_credited=date_credited
                )

                Notification.objects.create(
                    level="Low",
                    module="Finance",
                    notifier=request.user,
                    page="loans-allowances",
                    reciever=employee,
                    message="has awarded you for perfect attendance."
                )
            except EmployeeLogin.DoesNotExist:
                messages.error(request, f"Employee with ID {idnumber} does not exist.")
            except FinanceType.DoesNotExist:
                messages.error(request, f"Finance type '{finance_type_name}' not found.")
            except Exception as e:
                messages.error(request, f"Error processing employee {idnumber}: {e}")

        messages.success(request, "Perfect attendance uploaded successfully.")
        return redirect('admin-finance')

    return redirect('admin-finance')

@login_required(login_url='login')
@user_passes_test(is_admin, login_url='Dashboard:Dashboard')
def import_medicine(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect('admin-finance')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception as e:
            messages.error(request, f"Error loading file: {e}")
            return redirect('admin-finance')

        Medicine.objects.all().delete()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                idnumber = row[0]
                finance_type_name = row[1]
                started_at = row[2]
                ended_at = row[3]
                amount = row[4]
                claimed_amount = row[5]
                balance = row[6]

                employee = EmployeeLogin.objects.get(idnumber=idnumber)
                finance_type = FinanceType.objects.get(name=finance_type_name) if finance_type_name else None

                Medicine.objects.create(
                    employee=employee,
                    type=finance_type,
                    started_at=started_at,
                    ended_at=ended_at,
                    amount=amount,
                    claimed_amount=claimed_amount,
                    balance=balance
                )

                Notification.objects.create(
                    level="Low",
                    module="Finance",
                    notifier=request.user,
                    page="loans-allowances",
                    reciever=employee,
                    message="updated you medicine allowances."
                )
            except EmployeeLogin.DoesNotExist:
                messages.error(request, f"Employee with ID {idnumber} does not exist.")
            except FinanceType.DoesNotExist:
                messages.error(request, f"Finance type '{finance_type_name}' not found.")
            except Exception as e:
                messages.error(request, f"Error processing employee {idnumber}: {e}")

        messages.success(request, "Medicine allowances uploaded successfully.")
        return redirect('admin-finance')

    return redirect('admin-finance')


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='Dashboard:Dashboard')
def import_loans(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect('admin-finance')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception as e:
            messages.error(request, f"Error loading file: {e}")
            return redirect('admin-finance')

        Loan.objects.all().delete()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                idnumber = row[0]
                loan_type_name = row[1]
                date_started = row[2]
                loan_amount = row[3]
                payment = row[4]

                employee = EmployeeLogin.objects.get(idnumber=idnumber)
                loan_type = FinanceType.objects.get(name=loan_type_name) if loan_type_name else None

                Loan.objects.create(
                    employee=employee,
                    loan_type=loan_type,
                    date_started=date_started,
                    loan_amount=loan_amount,
                    payment=payment,
                )

                Notification.objects.create(
                    level="Low",
                    module="Finance",
                    notifier=request.user,
                    page="loans-allowances",
                    reciever=employee,
                    message="updated you loan balances."
                )
            except EmployeeLogin.DoesNotExist:
                messages.error(request, f"Employee with ID {idnumber} does not exist.")
            except FinanceType.DoesNotExist:
                messages.error(request, f"Finance type '{loan_type_name}' not found.")
            except Exception as e:
                messages.error(request, f"Error processing employee {idnumber}: {e}")

        messages.success(request, "Loans uploaded successfully.")
        return redirect('admin-finance')

    return redirect('admin-finance')

@login_required(login_url='login')
@user_passes_test(is_admin, login_url='Dashboard:Dashboard')
def import_savings(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect('admin-finance')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception as e:
            messages.error(request, f"Error loading file: {e}")
            return redirect('admin-finance')

        Savings.objects.all().delete()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                idnumber = row[0]
                finance_type_name = row[1]
                date_started = row[2]
                desired_savings = row[3]
                savings = row[4]

                employee = EmployeeLogin.objects.get(idnumber=idnumber)
                finance_type = FinanceType.objects.get(name=finance_type_name) if finance_type_name else None

                Savings.objects.create(
                    employee=employee,
                    savings_type=finance_type,
                    date_started=date_started,
                    desired_savings=desired_savings,
                    savings=savings,
                )

                Notification.objects.create(
                    level="Low",
                    module="Finance",
                    notifier=request.user,
                    page="loans-allowances",
                    reciever=employee,
                    message="updated you savings balances."
                )
            except EmployeeLogin.DoesNotExist:
                messages.error(request, f"Employee with ID {idnumber} does not exist.")
            except FinanceType.DoesNotExist:
                messages.error(request, f"Finance type '{finance_type_name}' not found.")
            except Exception as e:
                messages.error(request, f"Error processing employee {idnumber}: {e}")

        messages.success(request, "Saving uploaded successfully.")
        return redirect('admin-finance')

    return redirect('admin-finance')

@login_required(login_url='login')
def payslip_view(request):
    payslips = EmployeePayslip.objects.filter(employee=request.user).order_by('-uploaded_at')

    context={
        'payslips':payslips
    }
    return render(request, 'Accounting/payslip-view.html', context)

@login_required(login_url='login')
def admin_payslip_view(request, pk):
    employee = get_object_or_404(EmployeeLogin, id=pk)
    payslips = EmployeePayslip.objects.filter(employee=employee).order_by('-uploaded_at')

    if not payslips.exists():
        payslips = None
        
    context={
        # 'employee':employee,
        'payslips':payslips
    }
    return render(request, 'Accounting/payslip-view.html', context)

@login_required(login_url='login')
@user_passes_test(is_admin, login_url='Dashboard:Dashboard')
def upload_payslips(request):
    if request.method == 'POST':
        payroll_start = request.POST.get('payroll_start')
        payroll_end = request.POST.get('payroll_end')
        if not payroll_start:
            messages.error(request, "Please select a cut-off date")
            return redirect('admin-finance"')
        
        if not payroll_end:
            messages.error(request, "Please select a cut-off date")
            return redirect('admin-finance"')
        
        try:
            payroll_start = datetime.strptime(payroll_start, '%Y-%m-%d').date()
            payroll_end = datetime.strptime(payroll_end, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Invalid payroll date format")
            return redirect('admin-finance"')
        
        files = request.FILES.getlist('payslips')
        
        successful_uploads = 0
        failed_uploads = []

        for file in files:
            try:
                if file.size > 400 * 1024:  # 400KB
                    failed_uploads.append(f"{file.name}: File size exceeds 400KB")
                    continue              
                employee_id = os.path.splitext(file.name)[0]
                
                try:

                    employee = EmployeeLogin.objects.get(idnumber=employee_id)
                    payslip, created = EmployeePayslip.objects.update_or_create(
                        employee=employee,
                        payroll_start=payroll_start,
                        payroll_end=payroll_end,
                        defaults={'payslip': file}
                    )

                    Notification.objects.create(
                        level="Low",
                        module="Finance",
                        notifier=request.user,
                        page="payslip",
                        reciever=employee,
                        message="has uploaded your new payslip."
                    )
                    successful_uploads += 1
                
                except ObjectDoesNotExist:
                    failed_uploads.append(f"{file.name}: No employee found with ID {employee_id}")
                
            except Exception as e:
                failed_uploads.append(f"{file.name}: {str(e)}")
        
        if successful_uploads > 0:
            messages.success(request, f'Successfully uploaded {successful_uploads} payslips')
        
        if failed_uploads:
            messages.error(request, "Some uploads failed:\n" + "\n".join(failed_uploads))
        
        return redirect('admin-finance')
    
    return redirect('admin-finance')

@login_required(login_url='login')
def get_payslip_url(request, payslip_id):
    payslip = get_object_or_404(EmployeePayslip, id=payslip_id, employee=request.user)
    print(f"Payslip URL: {payslip.payslip.url}")
    return JsonResponse({
        'url': payslip.payslip.url
    })

@login_required(login_url='login')
def send_payslip_email(request, pk):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
        
    try:
        payslip = get_object_or_404(EmployeePayslip, employee=request.user, id=pk)
        personal_info = get_object_or_404(PersonalInformation, name=request.user)

        recipient_email = request.user.email
        if not recipient_email:
            messages.error(request, "No email address found for your account.")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', 'payslip'))

        if not payslip.payslip or not os.path.exists(payslip.payslip.path):
            messages.error(request, "Payslip file not found.")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', 'payslip'))

        # Generate password using birthdate
        birthdate = personal_info.birth_date.strftime("%m%d%Y")
        password = f"Repco{birthdate}"

        # Create encrypted PDF
        reader = PdfReader(payslip.payslip.path)
        writer = PdfWriter()

        # Add all pages to the writer
        for page in reader.pages:
            writer.add_page(page)

        # Encrypt the PDF
        writer.encrypt(password)

        # Save the encrypted PDF to a temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
            writer.write(temp_file)
            temp_file_path = temp_file.name
        
        # Prepare email
        email_subject = f'Payslip for {payslip.payroll_start.strftime("%B %d, %Y")} - {payslip.payroll_end.strftime("%B %d, %Y")}'
        email_body = f'''Dear {request.user.name},

Please find attached your payslip for {payslip.payroll_start.strftime("%B %d, %Y")} - {payslip.payroll_end.strftime("%B %d, %Y")}.

The PDF is password protected. Use your standard password format "RepcoMMDDYYYY" to open it.

Sincerely,
FAD Department
Ryonan Electric Philippines Corporation'''

        # Send email with encrypted PDF
        email = EmailMessage(
            subject=email_subject,
            body=email_body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[recipient_email]
        )
        email.attach_file(temp_file_path)
        email.send(fail_silently=False)

        # Clean up - delete temporary file
        os.unlink(temp_file_path)

        payslip.status = "Sent to Email"
        payslip.received_date = timezone.now()
        payslip.save()
        
        messages.success(request, f"Encrypted payslip sent successfully to {recipient_email}!")
        
    except EmployeePayslip.DoesNotExist:
        messages.error(request, "Selected payslip not found.")
    except SMTPException as e:
        messages.error(request, f"Email sending failed: {str(e)}")
    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {str(e)}")
    
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', 'payslip'))