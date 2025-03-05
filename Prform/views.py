from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import PRForm
from UserLogin.models import EmployeeLogin
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from django.utils.dateparse import parse_date
from Notification.models import Notification
from django.db.models import Q

@login_required(login_url='login')
def prf_view(request):
    prfs = PRForm.objects.filter(request_by=request.user).order_by('-submitted_at')
    types = PRForm.objects.values_list('request_type', flat=True).distinct()

    context={
        'prfs':prfs,
        'types':types
    }
    return render(request, 'Prform/Pr-form.html', context)


@login_required(login_url='login')
def submit_prf(request):
    prfs = PRForm.objects.filter(request_by=request.user).order_by('-submitted_at')
    types = PRForm.objects.values_list('request_type', flat=True).distinct()
    admins = EmployeeLogin.objects.filter(Q(is_accounting=True) | Q(is_hr=True))
    if request.method == "POST":
        request_type = request.POST.get('prfform')
        request_type_others = request.POST.get('prf_others')
        ctrl_no = request.POST.get('ctrl_no')
        status = "Pending"
        purpose = request.POST.get('purpose')

        if request_type == "Others":
            request_type = request_type_others

        PRFormForm = PRForm(
            request_by=request.user,
            request_type=request_type,
            ctrl_no=ctrl_no,
            status=status,
            purpose=purpose
        )

        for admin in admins:
                Notification.objects.create(
                    level="Low",
                    module='PRF',
                    notifier=request.user,
                    page="admin-prf",
                    reciever=admin,
                    message=f'has submitted a {request_type} request.'
                )

        PRFormForm.save()
        messages.success(request, 'PRF request submitted.')
        return redirect('user-prf')
    
    context={
        'prfs':prfs,
        'types':types
    }
    return render(request, 'Prform/Pr-form.html', context)

@login_required(login_url='login')
def edit_prf(request, pk):
    selectedPRF = get_object_or_404(PRForm, id=pk)

    if request.method == "POST":
        ctrl_no = request.POST.get('ctrl_no')
        purpose = request.POST.get('purpose')

        selectedPRF.ctrl_no=ctrl_no
        selectedPRF.purpose=purpose
        
        selectedPRF.save()
        messages.success(request, 'PRF submitted successfully.')
        return redirect('user-prf')
    
    context = {
        'selectedPRF':selectedPRF,
    }
    return render(request, 'Prform/Pr-form.html', context)

@login_required(login_url='login')
def cancel_prf(request, pk):
    cancelPRF = get_object_or_404(PRForm, id=pk)

    if request.method == "POST":
        status = 'Cancelled'
        cancelPRF.status=status
        
        cancelPRF.save()
        messages.success(request, 'PRF cancelled.')
        return redirect('user-prf')
    
    context = {
        'cancelPRF':cancelPRF,
    }
    return render(request, 'Prform/Pr-form.html', context)

@login_required(login_url='login')
def admin_prf(request):
    if not request.user.is_hr and not request.user.is_accounting:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    prfs = PRForm.objects.all().order_by('-submitted_at')
    types = PRForm.objects.values_list('request_type', flat=True).distinct()

    for prf in prfs:
        try:
            employee = EmployeeLogin.objects.get(idnumber=prf.request_by.idnumber)
            prf.employee_name = employee.name
        except EmployeeLogin.DoesNotExist:
            prf.employee_name = "Unknown" 

    context={
        'prfs':prfs,
        'types':types
    }
    return render(request, 'Prform/admin-prf.html', context)


@login_required(login_url='login')
def update_prf_status(request, pk, action):
    prf = get_object_or_404(PRForm, id=pk)

    if request.method == "POST":
        if action == 'approve':
            prf.status = 'Approved'
            prf.approved_by = request.user.name

            Notification.objects.create(
                level="Low",
                module="PRF",
                notifier=request.user,
                page="user-prf",
                reciever=prf.request_by,
                message=f'approved your {prf.request_type.lower()}.'
            )
            
            messages.success(request, 'PRF approved.')
        elif action == 'disapprove':
            prf.status = 'Disapproved'
            prf.approved_by = request.user.name
            messages.success(request, 'PRF disapproved.')

            Notification.objects.create(
                level="Low",
                module="PRF",
                notifier=request.user,
                page="user-prf",
                reciever=prf.request_by,
                message=f'disapproved your {prf.request_type.lower()}.'
            )
        
        prf.save()
        return redirect('admin-prf')
    
    context = {
        'prf': prf,
        'action': action,
    }
    return render(request, 'Prform/admin-prf.html', context)


@login_required(login_url='login')
def export_prf(request):
    status_filter = request.GET.get('status', 'All')
    type_filter = request.GET.get('type', 'All')
    from_date = request.GET.get('from_date', None)
    to_date = request.GET.get('to_date', None)
    
    # Base queryset
    prfs = PRForm.objects.all().order_by('-submitted_at')
    
    # Apply status filter if not 'All'
    if status_filter and status_filter != 'All':
        prfs = prfs.filter(status=status_filter)
    
    # Apply type filter if not 'All'
    if type_filter and type_filter != 'All':
        prfs = prfs.filter(request_type=type_filter)
    
    # Apply date range filter if both dates are provided
    if from_date:
        from_date = parse_date(from_date)
        prfs = prfs.filter(submitted_at__gte=from_date)
    if to_date:
        to_date = parse_date(to_date)
        prfs = prfs.filter(submitted_at__lte=to_date)

    wb = Workbook()
    ws = wb.active

    ws['A1'] = "Ryonan Electric Philippines Corporation"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = "Monthly PRF Report"
    
    headers = ['Date Filed', 'Reference Number', 'ID Number', 'Name', 'Request Type', 'Control Number', 'Purpose', 'Status']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)

    for row, prf in enumerate(prfs, start=4):
        ws.cell(row=row, column=1, value=prf.submitted_at.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=f'PRF{prf.id:04d}')
        ws.cell(row=row, column=3, value=prf.request_by.idnumber)
        ws.cell(row=row, column=4, value=prf.request_by.name)
        ws.cell(row=row, column=5, value=prf.request_type)
        ws.cell(row=row, column=6, value=prf.ctrl_no)
        ws.cell(row=row, column=7, value=prf.purpose)
        ws.cell(row=row, column=8, value=prf.status)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=PRF_Report.xlsx'

    wb.save(response)
    return response