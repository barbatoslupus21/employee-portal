import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import logout
from UserLogin.models import EmployeeLogin
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import PersonalInformation, EmploymentInformation
from Training.models import ParticipantResponse
from Settings.models import Gender, Lines, Department, Position, Status
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from django.http import HttpResponse
from Notification.models import Notification
from django.contrib.auth import update_session_auth_hash
import re
from Prform.models import PRForm
from django.db.models import Prefetch, F, Count, FloatField
from django.db.models.functions import Cast
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from Certificate.models import Awardees
from rest_framework.permissions import IsAuthenticated
from Overtime.models import PickUpPoint
from django.contrib.auth.hashers import make_password
from django.core.validators import validate_email
from django.core.exceptions import ValidationError

@login_required(login_url='login')
def profile_view(request):
    certificates = Awardees.objects.filter(awardeee=request.user).order_by('-posted_at')[:5]
    certificate_count = Awardees.objects.filter(awardeee=request.user).count()
    extra_certificate_count = certificate_count - 5 if certificate_count > 5 else None

    trainings = ParticipantResponse.objects.filter(employee=request.user).order_by('-posted_at')[:5]
    trainingCount = ParticipantResponse.objects.filter(employee=request.user).count()
    extra_training_count = trainingCount - 5 if trainingCount > 5 else None

    try:
        PersonalInfo = PersonalInformation.objects.get(name=request.user)
    except PersonalInformation.DoesNotExist:
        PersonalInfo = None

    try:
        EmpInfo = EmploymentInformation.objects.get(name=request.user)
    except EmploymentInformation.DoesNotExist:
        EmpInfo = None
    context = {
        'PersonalInfo':PersonalInfo,
        'EmpInfo':EmpInfo,
        'trainings':trainings,
        'trainingCount':trainingCount,
        'extra_training_count':extra_training_count,
        'certificates':certificates,
        'certificate_count':certificate_count,
        'extra_certificate_count':extra_certificate_count
    }
    return render(request, 'Profile/user-profile.html', context)

@login_required(login_url='login')
def update_profile(request):
    try:
        personalInfo = PersonalInformation.objects.get(name=request.user)
    except PersonalInformation.DoesNotExist:
        personalInfo = None

    try:
        empInfo = EmploymentInformation.objects.get(name=request.user)
    except EmploymentInformation.DoesNotExist:
        empInfo = None

    genders = Gender.objects.all()
    lines = Lines.objects.all()
    departments = Department.objects.all()
    leaders = EmployeeLogin.objects.filter(is_approver=True).exclude(is_admin=True)
    pick_ups = PickUpPoint.objects.all()
    context = {
        'personalInfo':personalInfo,
        'empInfo':empInfo,
        'genders':genders,
        'lines':lines,
        'departments':departments,
        'leaders':leaders,
        'pick_ups':pick_ups
    }
    
    return render(request, 'Profile/update-profile.html', context)


@login_required(login_url='login')
def save_profile(request):
    LoggedUser= request.user
    try:
        personal_info = PersonalInformation.objects.get(name=request.user)
    except PersonalInformation.DoesNotExist:
        personal_info = None

    try:
        employment_info = EmploymentInformation.objects.get(name=request.user)
    except EmploymentInformation.DoesNotExist:
        employment_info = None

    if request.method == "POST":
        # Retrieve form data
        avatar = request.FILES.get('user_avatar')
        firstname = request.POST.get('first_name')
        lastname = request.POST.get('last_name')
        personal_email = request.POST.get('personal_email')
        
        middlename = request.POST.get('middle_name')  
        nickname = request.POST.get('username')
        gender = request.POST.get('gender')
        user_gender = Gender.objects.get(name=gender)
        line = request.POST.get('line')
        user_line = Lines.objects.get(line=line)
        department = request.POST.get('department')
        user_department = Department.objects.get(abreviation=department)
        birth_place = request.POST.get('birthplace')
        birth_date = request.POST.get('birthdate') 
        work_email = request.POST.get('work_email')
        contact_number = request.POST.get('contact_number')
        mother = request.POST.get('mother')
        father = request.POST.get('father')

        # present address
        present_street = request.POST.get('present_street')
        present_brgy = request.POST.get('present_baranggay')
        present_city = request.POST.get('present_city')
        present_province = request.POST.get('present_province')

        # provincial address
        provincial_street = request.POST.get('provincial_street')
        provincial_brgy = request.POST.get('provincial_baranggay')
        provincial_city = request.POST.get('provincial_city')
        provincial_province = request.POST.get('provincial_province')

        # approver
        approver = request.POST.get('supervisor')
        user_approver = EmployeeLogin.objects.get(id=approver)

        # shuttle
        shuttle = request.POST.get('shuttle')
        user_shuttle = PickUpPoint.objects.get(id=shuttle)

        # emergency contact
        emergency_first = request.POST.get('contact_first')
        emergency_middle = request.POST.get('contact_middle')
        emergency_last = request.POST.get('contact_last')
        emergency_relation = request.POST.get('emergency_relation')
        emergency_contact = request.POST.get('emergency_number')

        emergency_street = request.POST.get('emergency_street')
        emergency_brgy = request.POST.get('emergency_baranggay')
        emergency_city = request.POST.get('emergency_city')
        emergency_province = request.POST.get('emergency_province')

        # other info
        primary_school = request.POST.get('Primary')
        secondary_school = request.POST.get('Secondary')
        vocational_school = request.POST.get('Vocational')
        tertiary_school = request.POST.get('Tertiary')
        spouse = request.POST.get('spouse_name')
        no_of_children = request.POST.get('no_of_children')
        children = request.POST.get('children_name')

        if personal_info:
            personal_info.middlename=middlename
            personal_info.nickname=nickname
            personal_info.gender=user_gender
            personal_info.birth_date=birth_date
            personal_info.birth_place=birth_place
            personal_info.contact_number=contact_number
            personal_info.mother=mother
            personal_info.father=father
            personal_info.work_email=work_email
            personal_info.present_street=present_street
            personal_info.present_baranggay=present_brgy
            personal_info.present_city=present_city
            personal_info.present_province=present_province
            personal_info.provincial_street=provincial_street
            personal_info.provincial_baranggay=provincial_brgy
            personal_info.provincial_city=provincial_city
            personal_info.provincial_province=provincial_province
            personal_info.contact_firstname=emergency_first
            personal_info.contact_middlename=emergency_middle
            personal_info.contact_lastname=emergency_last
            personal_info.contact_relation=emergency_relation
            personal_info.contact_no=emergency_contact
            personal_info.contact_street=emergency_street
            personal_info.contact_baranggay=emergency_brgy
            personal_info.contact_city=emergency_city
            personal_info.contact_province=emergency_province
            personal_info.primary_school=primary_school
            personal_info.secondary_school=secondary_school
            personal_info.vocational_school=vocational_school
            personal_info.tertiary_school=tertiary_school
            personal_info.spouse=spouse
            personal_info.no_of_children=no_of_children
            personal_info.children=children

        else:
            personal_info = PersonalInformation(
                name = LoggedUser,
                middlename=middlename,
                nickname=nickname,
                gender=user_gender,
                birth_date=birth_date,
                birth_place=birth_place,
                contact_number=contact_number,
                mother=mother,
                father=father,
                work_email=work_email,
                present_street=present_street,
                present_baranggay=present_brgy,
                present_city=present_city,
                present_province=present_province,
                provincial_street=provincial_street,
                provincial_baranggay=provincial_brgy,
                provincial_city=provincial_city,
                provincial_province=provincial_province,
                contact_firstname=emergency_first,
                contact_middlename=emergency_middle,
                contact_lastname=emergency_last,
                contact_relation=emergency_relation,
                contact_no=emergency_contact,
                contact_street=emergency_street,
                contact_baranggay=emergency_brgy,
                contact_city=emergency_city,
                contact_province=emergency_province,
                primary_school=primary_school,
                secondary_school=secondary_school,
                vocational_school=vocational_school,
                tertiary_school=tertiary_school,
                spouse=spouse,
                no_of_children=no_of_children,
                children=children
            )

        if employment_info:
            employment_info.approver=user_approver
            employment_info.line=user_line
            employment_info.shuttle=user_shuttle
            employment_info.department=user_department
        else:
            employment_info = EmploymentInformation(
                name=LoggedUser,
                approver=user_approver,
                shuttle=user_shuttle,
                line=user_line,
                department=user_department
            )

        if avatar:
            LoggedUser.avatar=avatar
            LoggedUser.firstname=firstname
            LoggedUser.lastname=lastname
            LoggedUser.name=f'{firstname} {lastname}'
            LoggedUser.email=personal_email

        employment_info.save()
        personal_info.save()
        LoggedUser.save()
        messages.success(request, 'Profile updated sucessfully.')
        return redirect('user-profile')
    
    return redirect('user-profile')

@login_required(login_url='login')
def employee_accounts(request):
    if not request.user.is_hr:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    accounts = EmployeeLogin.objects.all().exclude(is_admin=True)
    statuses = Status.objects.all()
    lines = Lines.objects.all()

    for account in accounts:
        try:
            participant_info = EmploymentInformation.objects.filter(name=account).first()

            if participant_info:
                account.status = participant_info.Status.name if participant_info.Status else None
                account.line = participant_info.line.line if participant_info.line else None
            else:
                account.status = None
                account.line = None
        except EmploymentInformation.DoesNotExist:
            account.status = None
            account.line = None

    context = {
        'accounts': accounts,
        'statuses': statuses,
        'lines': lines,
    }
    return render(request, 'Profile/admin-account.html', context)

@login_required(login_url='login')
def approval(request, pk, action):
    account = get_object_or_404(EmployeeLogin, id=pk)

    if request.method == "POST":
        if action == 'approve':
            account.is_approved = "Approved"
            messages.success(request, 'Account approved.')
        elif action == 'disapprove':
            account.is_approved = "Disapproved"
            account.is_active = False
            messages.success(request, 'Account disapproved.')
        
        account.save()
        return redirect('accounts')

    return redirect('accounts')

@login_required(login_url='login')
def reset_password(request, pk): 
    account = get_object_or_404(EmployeeLogin, id=pk) 
    reciever = EmployeeLogin.objects.get(id=pk)
    if request.method == 'POST':
        account.set_password('12345') 

        Notification.objects.create(
            level="Medium",
            module="Account",
            notifier=request.user,
            page="account-settings",
            reciever=reciever,
            message="has reset your password."
        )

        account.save()
        messages.success(request, 'Account Reset.') 
        return redirect('accounts')    
    return redirect('accounts')

@login_required(login_url='login')
def lock_account(request, pk): 
    account = get_object_or_404(EmployeeLogin, id=pk) 
    reciever = EmployeeLogin.objects.get(id=pk)
    if request.method == 'POST':
        account.is_locked = True
        account.save()
        messages.success(request, 'Account locked.') 
        return redirect('accounts')      
    return redirect('accounts')

@login_required(login_url='login')
def unlock_account(request, pk): 
    account = get_object_or_404(EmployeeLogin, id=pk) 
    reciever = EmployeeLogin.objects.get(id=pk)
    if request.method == 'POST':
        account.is_locked=False 

        Notification.objects.create(
            level="Medium",
            module="Account",
            notifier=request.user,
            page="Dashboard:Dashboard",
            reciever=reciever,
            message="has unlock your account."
        )

        account.save()
        messages.success(request, 'Account unlocked.') 
        return redirect('accounts')    
    return redirect('accounts')

@login_required(login_url='login')
def account_view(request, pk):
    trainings = ParticipantResponse.objects.filter(employee=pk).order_by('-posted_at')[:5]
    trainingCount = ParticipantResponse.objects.filter(employee=pk).count()
    extra_training_count = trainingCount - 5 if trainingCount > 5 else None
    selectedAccount = EmployeeLogin.objects.get(id=pk)
    try:
        PersonalInfo = PersonalInformation.objects.get(name=pk)
    except PersonalInformation.DoesNotExist:
        messages.error(request, "The selected user hasn't filled out their personal information.")
        PersonalInfo = None
    
    try:
        EmpInfo = EmploymentInformation.objects.get(name=pk)
    except EmploymentInformation.DoesNotExist:
        messages.error(request, "The selected user hasn't filled out their employment information.")
        return redirect('accounts')
    
    approvers = EmployeeLogin.objects.filter(is_approver=True).exclude(is_admin=True)
    
    if EmpInfo.approver:
        approvers = approvers.exclude(name=EmpInfo.approver.name)

    positions = Position.objects.all()
    if EmpInfo.position:
        positions = positions.exclude(id=EmpInfo.position.id)

    statuses = Status.objects.all()
    if EmpInfo.Status:
        statuses = statuses.exclude(id=EmpInfo.Status.id)

    departments = Department.objects.all()
    if EmpInfo.department:
       departments = departments.exclude(id=EmpInfo.department.id)

    lines = Lines.objects.all()
    if EmpInfo.line:
        lines = lines.exclude(id=EmpInfo.line.id)

    context = {
        'PersonalInfo':PersonalInfo,
        'EmpInfo':EmpInfo,
        'trainings':trainings,
        'trainingCount':trainingCount,
        'extra_training_count':extra_training_count,
        'selectedAccount':selectedAccount,
        'positions':positions,
        'statuses':statuses,
        'departments':departments,
        'lines':lines,
        'approvers':approvers
    }
    return render(request, 'Profile/admin-account-view.html', context)

@login_required(login_url='login')
def edit_account(request, pk):

    selectedUser = get_object_or_404(EmployeeLogin, id=pk)
    try:
        selectedAccount = EmploymentInformation.objects.get(name=pk)
    except EmploymentInformation.DoesNotExist:
        selectedAccount = None

    if request.method == "POST":
        idnumber = request.POST.get('idnumber')
        position = request.POST.get('position')
        account_position = Position.objects.get(id=position)

        line = request.POST.get('line')
        account_line = Lines.objects.get(id=line)

        department = request.POST.get('department')
        account_department = Department.objects.get(id=department)

        status = request.POST.get('status')
        account_status = Status.objects.get(id=status)

        type = request.POST.get('type')
        date_hired = request.POST.get('date_hired')
        tin_number = request.POST.get('tin_number')
        sss_number = request.POST.get('sss_number')
        hdmf_number = request.POST.get('hdmf_number')
        philhealth = request.POST.get('philhealth')
        bank_account = request.POST.get('bank_account')

        is_approver = request.POST.get('is_approver')

        if selectedAccount:
            selectedAccount.position=account_position
            selectedAccount.line=account_line
            selectedAccount.department=account_department
            selectedAccount.Status=account_status
            selectedAccount.type=type
            selectedAccount.date_hired=date_hired
            selectedAccount.tin_number=tin_number
            selectedAccount.sss_number=sss_number
            selectedAccount.hdmf_number=hdmf_number
            selectedAccount.philhealth=philhealth
            selectedAccount.bank_account=bank_account
        else:
            selectedAccount = EmploymentInformation(
                name=pk,
                position=account_position,
                line=account_line,
                department=account_department,
                Status=account_status,
                type=type,
                date_hired=date_hired,
                tin_number=tin_number,
                sss_number=sss_number,
                hdmf_number=hdmf_number,
                philhealth=philhealth,
                bank_account=bank_account
            )
        if is_approver == "True":
            selectedUser.is_approver = True
        else:
            selectedUser.is_approver = False
        selectedUser.idnumber=idnumber
        selectedUser.save()
        selectedAccount.save()
        messages.success(request, 'Employment information updated.')
        return redirect('accounts')
    
    return redirect('accounts')

@login_required(login_url='login')
def employment_info_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect('accounts')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception as e:
            messages.error(request, f"Error loading file: {e}")
            return redirect('accounts')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                idnumber = row[0]
                position_name = row[1]
                line_name = row[2]
                department_name = row[3]
                status_name = row[4]
                emp_type = row[5]
                date_hired = row[6]
                tin_number = row[7]
                sss_number = row[8]
                hdmf_number = row[9]
                philhealth = row[10]
                bank_account = row[11]

                employee = EmployeeLogin.objects.get(idnumber=idnumber)
                line = Lines.objects.get(line=line_name) if line_name else None
                department = Department.objects.get(abreviation=department_name) if department_name else None
                position = Position.objects.get(name=position_name) if position_name else None
                status = Status.objects.get(name=status_name) if status_name else None

                if EmploymentInformation.objects.filter(name=employee).exists():
                    messages.warning(request, f"Employee with ID {idnumber} already has employment information.")
                    continue 

                EmploymentInformation.objects.create(
                    name=employee,
                    position=position,
                    line=line,
                    department=department,
                    Status=status,
                    type=emp_type,
                    date_hired=date_hired,
                    tin_number=tin_number,
                    sss_number=sss_number,
                    hdmf_number=hdmf_number,
                    philhealth=philhealth,
                    bank_account=bank_account
                )
            except Exception as e:
                messages.error(request, f"Error processing employee {idnumber}: {e}")

        messages.success(request, "Employment information successfully imported.")
        return redirect('accounts')

    return redirect('accounts')

@login_required(login_url='login')
def personal_info_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect('accounts')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception as e:
            messages.error(request, f"Error loading file: {e}")
            return redirect('accounts')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                idnumber = row[0]
                middlename = row[1]
                nickname = row[2]
                user_gender = row[3]
                birth_date = row[4]
                birth_place = row[5]
                contact_number = row[6]
                mother = row[7]
                father = row[8]
                work_email = row[9]
                present_street = row[10]
                present_baranggay = row[11]
                present_city = row[12]
                present_province = row[13]
                provincial_street = row[14]
                provincial_baranggay = row[15]
                provincial_city = row[16]
                provincial_province = row[17]
                contact_firstname = row[18]
                contact_middlename = row[19]
                contact_lastname = row[20]
                contact_relation = row[21]
                contact_no = row[22]
                contact_street = row[23]
                contact_baranggay = row[24]
                contact_city = row[25]
                contact_province = row[26]
                primary_school = row[27]
                secondary_school = row[28]
                vocational_school = row[29]
                tertiary_school = row[30]
                spouse = row[31]
                no_of_children = row[32]
                children = row[33]

                employee = EmployeeLogin.objects.get(idnumber=idnumber)
                gender = Gender.objects.get(name=user_gender) if user_gender else None

                if PersonalInformation.objects.filter(name=employee).exists():
                    messages.warning(request, f"Employee with ID {idnumber} already has personal information.")
                    continue 

                PersonalInformation.objects.create(
                    name = employee,
                    middlename = middlename,
                    nickname = nickname,
                    gender = gender,
                    birth_date = birth_date,
                    birth_place = birth_place,
                    contact_number = contact_number,
                    mother = mother,
                    father = father,
                    work_email = work_email,

                    present_street = present_street,
                    present_baranggay = present_baranggay,
                    present_city = present_city,
                    present_province = present_province,
                    
                    provincial_street = provincial_street,
                    provincial_baranggay = provincial_baranggay,
                    provincial_city = provincial_city,
                    provincial_province = provincial_province,
                    contact_firstname = contact_firstname,
                    contact_middlename = contact_middlename,
                    contact_lastname = contact_lastname,
                    contact_relation = contact_relation,
                    contact_no = contact_no,
                    contact_street = contact_street,
                    contact_baranggay = contact_baranggay,
                    contact_city = contact_city,
                    contact_province = contact_province,

                    primary_school = primary_school,
                    secondary_school = secondary_school,
                    vocational_school = vocational_school,
                    tertiary_school = tertiary_school,
                    spouse = spouse,
                    no_of_children = no_of_children,
                    children = children,
                )
            except Exception as e:
                messages.error(request, f"Error processing employee {idnumber}: {e}")

        messages.success(request, "Personal information successfully imported.")
        return redirect('accounts')

    return redirect('accounts')

@login_required(login_url='login')
def export_excel(request):
    wb = Workbook()
    
    # personal Sheet
    ws = wb.active
    ws.title = "Personal Informations"
    ws['A1'] = "Ryonan Electric Philippines Corporation"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = "Employee Personal Information Masterlist"
    ws['A4'] = "ID number"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = "Name"
    ws['B4'].font = Font(bold=True)
    ws['C4'] = "Nickname"
    ws['C4'].font = Font(bold=True)
    ws['D4'] = "Gender"
    ws['D4'].font = Font(bold=True)
    ws['E4'] = "Place of Birth"
    ws['E4'].font = Font(bold=True)
    ws['F4'] = "Date of Birth"
    ws['F4'].font = Font(bold=True)
    ws['G4'] = "Contact Number"
    ws['G4'].font = Font(bold=True)
    ws['H4'] = "Mother's Name"
    ws['H4'].font = Font(bold=True)
    ws['I4'] = "Father's Name"
    ws['I4'].font = Font(bold=True)
    ws['J4'] = "Personal Email"
    ws['J4'].font = Font(bold=True)
    ws['K4'] = "Work Email"
    ws['K4'].font = Font(bold=True)
    ws['L4'] = "Present Address"
    ws['L4'].font = Font(bold=True)
    ws['M4'] = "Provincial Address"
    ws['M4'].font = Font(bold=True)
    ws['N4'] = "Contact Person"
    ws['N4'].font = Font(bold=True)
    ws['O4'] = "Relation"
    ws['O4'].font = Font(bold=True)
    ws['P4'] = "Contact Number"
    ws['P4'].font = Font(bold=True)
    ws['Q4'] = "Address"
    ws['Q4'].font = Font(bold=True)
    ws['R4'] = "Primary School"
    ws['R4'].font = Font(bold=True)
    ws['S4'] = "Secondary School"
    ws['S4'].font = Font(bold=True)
    ws['T4'] = "Vocational School"
    ws['T4'].font = Font(bold=True)
    ws['U4'] = "Tertiary School"
    ws['U4'].font = Font(bold=True)
    ws['V4'] = "Spouse"
    ws['V4'].font = Font(bold=True)
    ws['W4'] = "Number of Children"
    ws['W4'].font = Font(bold=True)
    ws['X4'] = "Children Names"
    ws['X4'].font = Font(bold=True)
    
    personal_infos = PersonalInformation.objects.all().exclude(name__is_admin=True)

    for i, personal_info in enumerate(personal_infos, start=5):
        present_address = f'{personal_info.present_street}, {personal_info.present_baranggay}, {personal_info.present_city}, {personal_info.present_province} '
        provincial_address = f'{personal_info.provincial_street}, {personal_info.provincial_baranggay}, {personal_info.provincial_city}, {personal_info.provincial_province}'
        contact_person = f'{personal_info.contact_firstname} {personal_info.contact_middlename} {personal_info.contact_lastname}'
        contact_address = f'{personal_info.contact_street}, {personal_info.contact_baranggay}, {personal_info.contact_city}, {personal_info.contact_province}'

        ws[f'A{i}'] = personal_info.name.idnumber if personal_info.name.idnumber else ''
        ws[f'B{i}'] = personal_info.name.name if personal_info.name.name else ''
        ws[f'C{i}'] = personal_info.nickname if personal_info.nickname else ''
        ws[f'D{i}'] = personal_info.gender.name if personal_info.gender.name else ''
        ws[f'E{i}'] = personal_info.birth_place if personal_info.birth_place else ''
        ws[f'F{i}'] = personal_info.birth_date.strftime('%m/%d/%Y') if personal_info.birth_date else ''
        ws[f'G{i}'] = personal_info.contact_number if personal_info.contact_number else ''
        ws[f'H{i}'] = personal_info.mother if personal_info.mother else ''
        ws[f'I{i}'] = personal_info.father if personal_info.father else ''
        ws[f'J{i}'] = personal_info.name.email if personal_info.name.email else ''
        ws[f'K{i}'] = personal_info.work_email if personal_info.work_email else ''
        ws[f'L{i}'] = present_address if present_address else ''
        ws[f'M{i}'] = provincial_address if provincial_address else ''
        ws[f'N{i}'] = contact_person if contact_person else ''
        ws[f'O{i}'] = personal_info.contact_relation if personal_info.contact_relation else ''
        ws[f'P{i}'] = personal_info.contact_no if personal_info.contact_no else ''
        ws[f'Q{i}'] = contact_address if contact_address else ''
        ws[f'R{i}'] = personal_info.primary_school if personal_info.primary_school else ''
        ws[f'S{i}'] = personal_info.secondary_school if personal_info.secondary_school else ''
        ws[f'T{i}'] = personal_info.vocational_school if personal_info.vocational_school else ''
        ws[f'U{i}'] = personal_info.tertiary_school if personal_info.tertiary_school else ''
        ws[f'V{i}'] = personal_info.spouse if personal_info.spouse else ''
        ws[f'W{i}'] = personal_info.no_of_children if personal_info.no_of_children else ''
        ws[f'X{i}'] = personal_info.children if personal_info.children else ''     
    
    for row in ws['A4:X' + str(4 + len(personal_infos) + 1)]:
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # employment Sheet
    ws = wb.create_sheet("Employment Informations")
    ws['A1'] = "Ryonan Electric Philippines Corporation"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = "Employee Employment Information Masterlist"
    ws['A4'] = "ID number"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = "Name"
    ws['B4'].font = Font(bold=True)
    ws['C4'] = "Department"
    ws['C4'].font = Font(bold=True)
    ws['D4'] = "Line"
    ws['D4'].font = Font(bold=True)
    ws['E4'] = "Position"
    ws['E4'].font = Font(bold=True)
    ws['F4'] = "Status"
    ws['F4'].font = Font(bold=True)
    ws['G4'] = "Type"
    ws['G4'].font = Font(bold=True)
    ws['H4'] = "Date Hired"
    ws['H4'].font = Font(bold=True)
    ws['I4'] = "TIN Number"
    ws['I4'].font = Font(bold=True)
    ws['J4'] = "SSS Number"
    ws['J4'].font = Font(bold=True)
    ws['K4'] = "HDMF Number"
    ws['K4'].font = Font(bold=True)
    ws['L4'] = "PHILHEALTH"
    ws['L4'].font = Font(bold=True)
    ws['M4'] = "BanK Account"
    ws['M4'].font = Font(bold=True)
    ws['N4'] = "Supervisor"
    ws['N4'].font = Font(bold=True)
    
    employment_infos = EmploymentInformation.objects.all().exclude(name__is_admin=True)

    for i, employment_info in enumerate(employment_infos, start=5):
        ws[f'A{i}'] = employment_info.name.idnumber if employment_info.name.idnumber else ''
        ws[f'B{i}'] = employment_info.name.name if employment_info.name.name else ''
        ws[f'C{i}'] = employment_info.department.description if employment_info.department else ''
        ws[f'D{i}'] = employment_info.line.line if employment_info.line else ''
        ws[f'E{i}'] = employment_info.position.name if employment_info.position else ''
        ws[f'F{i}'] = employment_info.Status.name if employment_info.Status else ''
        ws[f'G{i}'] = employment_info.type if employment_info.type else ''
        ws[f'H{i}'] = employment_info.date_hired.strftime('%m/%Y') if employment_info.date_hired else ''
        ws[f'I{i}'] = employment_info.tin_number if employment_info.tin_number else ''
        ws[f'J{i}'] = employment_info.sss_number if employment_info.sss_number else ''
        ws[f'K{i}'] = employment_info.hdmf_number if employment_info.hdmf_number else ''
        ws[f'L{i}'] = employment_info.philhealth if employment_info.philhealth else ''
        ws[f'M{i}'] = employment_info.bank_account if employment_info.bank_account else ''
        ws[f'N{i}'] = employment_info.approver.name if employment_info.approver else ''
    
    for row in ws['A4:N' + str(4 + len(employment_infos) + 1)]:
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
     
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Employee_Information_Masterlist.xlsx'
    wb.save(response)
    
    return response

@login_required(login_url='login')
def account_setting(request):
    return render(request, 'Profile/account-settings.html')

@login_required(login_url='login')
def deactivate_account(request, pk): 
    account = get_object_or_404(EmployeeLogin, id=pk) 
    reciever = EmployeeLogin.objects.get(name="ADMIN") 
    if request.method == 'POST':
        Notification.objects.create(
            level="Medium",
            module="Account",
            notifier=request.user,
            page="accounts",
            reciever=reciever,
            message="has deleted their account."
        )
        account.delete() 
        messages.success(request, 'Account deleted.') 

        logout(request)
        return redirect('login')
        
    return redirect('login')

@login_required(login_url='login')
def change_password(request):
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        user = request.user
        
        # Server-side validation
        if not user.check_password(old_password):
            messages.error(request, 'Old password is incorrect.')
            return redirect('change-password')
        
        if new_password != confirm_password:
            messages.error(request, 'New passwords do not match.')
            return redirect('change-password')
        
        # Password criteria validation
        if len(new_password) < 12:
            messages.error(request, 'Password must be at least 12 characters long.')
            return redirect('change-password')
        
        if not re.search(r'[A-Z]', new_password):
            messages.error(request, 'Password must contain at least one uppercase letter.')
            return redirect('change-password')
        
        if not re.search(r'[a-z]', new_password):
            messages.error(request, 'Password must contain at least one lowercase letter.')
            return redirect('change-password')
        
        if not re.search(r'[!@#$%^&*(),_.?":{}|<>]', new_password):
            messages.error(request, 'Password must contain at least one special character.')
            return redirect('change-password')
        
        if not re.search(r'\d', new_password):
            messages.error(request, 'Password must contain at least one number.')
            return redirect('change-password')
        
        # Change password
        try:
            user.set_password(new_password)
            user.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password has been successfully changed.')
            return redirect('account-settings') 
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')
    
    return render(request, 'Profile/account-settings.html')

@login_required(login_url='login')
def id_request(request):
    if not request.user.is_hr:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    requestors = PRForm.objects.filter(request_type="ID Replacement")
    for requestor in requestors:
        try:
            request_department = EmploymentInformation.objects.filter(name=requestor.request_by.id).first()
            if request_department:
                requestor.department = request_department.department.abreviation
            else:
                requestor.department = None
                
            request_line = EmploymentInformation.objects.filter(name=requestor.request_by).first()
            if request_line:
                requestor.type = request_line.Status.name
            else:
                requestor.type = None
        except EmploymentInformation.DoesNotExist:
            request.department = None
            request.line = None

    context={
        'requestors':requestors
    }
    return render(request, 'Profile/id-request.html', context)

@login_required(login_url='login')
def id_request_status(request, pk, action):
    prf = get_object_or_404(PRForm, id=pk)

    if request.method == "POST":
        if action == 'approve':
            prf.status = 'Approved'
            prf.approved_by = request.user.name
            messages.success(request, 'PRF approved.')
        elif action == 'disapprove':
            prf.status = 'Disapproved'
            prf.approved_by = request.user.name
            messages.success(request, 'PRF disapproved.')
        
        prf.save()
        return redirect('id-request')
    
    context = {
        'prf': prf,
        'action': action,
    }
    return render(request, 'Prform/id-request.html', context)

@login_required(login_url='login')
def export_id_request(request):
    wb = Workbook()
    
    # personal Sheet
    ws = wb.active
    ws.title = "ID Replacement Request"
    ws['A1'] = "Ryonan Electric Philippines Corporation"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = "ID Replacement Request"
    ws['A4'] = "ID number"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = "Name"
    ws['B4'].font = Font(bold=True)
    ws['C4'] = "Nickname"
    ws['C4'].font = Font(bold=True)
    ws['D4'] = "Department"
    ws['D4'].font = Font(bold=True)
    ws['E4'] = "Position"
    ws['E4'].font = Font(bold=True)
    ws['F4'] = "Contact Person"
    ws['F4'].font = Font(bold=True)
    ws['G4'] = "Contact Person Address"
    ws['G4'].font = Font(bold=True)
    ws['H4'] = "Contact Person Number"
    ws['H4'].font = Font(bold=True)
    ws['I4'] = "SSS Number"
    ws['I4'].font = Font(bold=True)
    ws['J4'] = "TIN Number"
    ws['J4'].font = Font(bold=True)
    ws['K4'] = "Date Hired"
    ws['K4'].font = Font(bold=True)
    
    requestor = PRForm.objects.filter(request_type="ID Replacement").exclude(status="Disapproved").exclude(status="Approved")
    requestor_info = PersonalInformation.objects.filter(name__in=requestor.values_list('request_by', flat=True)).prefetch_related('name')
    emp_infos = EmploymentInformation.objects.filter(name__in=requestor.values_list('request_by', flat=True)).prefetch_related('name')

    personal_info_map = {info.name: info for info in requestor_info}
    emp_info_map = {info.name: info for info in emp_infos}

    for i, req in enumerate(requestor, start=5):
        personal_info = personal_info_map.get(req.request_by)
        emp_info = emp_info_map.get(req.request_by)

        if personal_info:
            contact_person = f'{personal_info.contact_firstname} {personal_info.contact_middlename} {personal_info.contact_lastname}'
            contact_address = f'{personal_info.contact_street}, {personal_info.contact_baranggay}, {personal_info.contact_city}, {personal_info.contact_province}'

            ws[f'A{i}'] = personal_info.name.idnumber if personal_info.name.idnumber else ''
            ws[f'B{i}'] = personal_info.name.name if personal_info.name.name else ''
            ws[f'C{i}'] = personal_info.nickname if personal_info.nickname else ''
            ws[f'D{i}'] = emp_info.department.abreviation if emp_info and emp_info.department else ''
            ws[f'E{i}'] = emp_info.position.name if emp_info and emp_info.position else ''
            ws[f'F{i}'] = contact_person if contact_person else ''
            ws[f'G{i}'] = contact_address if contact_address else ''
            ws[f'H{i}'] = personal_info.contact_no if personal_info.contact_no else ''
            ws[f'I{i}'] = emp_info.sss_number if emp_info else ''
            ws[f'J{i}'] = emp_info.tin_number if emp_info else ''
            ws[f'K{i}'] = emp_info.date_hired.strftime('%m/%d/%Y') if emp_info and emp_info.date_hired else ''
    
    for row in ws['A4:K' + str(4 + len(requestor) + 1)]:
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
     
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ID_Request.xlsx'
    wb.save(response)
    
    return response

@login_required(login_url='login')
def export_all_employees(request):
    wb = Workbook()
    
    ws = wb.active
    ws.title = "ID Request"
    
    headers = [
        ("Ryonan Electric Philippines Corporation", 1, True),
        ("Employee's ID Replacement", 2, False),
        ("ID number", 4, True),
        ("Name", 4, True),
        ("Nickname", 4, True),
        ("Department", 4, True),
        ("Position", 4, True),
        ("Contact Person", 4, True),
        ("Contact Person Address", 4, True),
        ("Contact Person Number", 4, True),
        ("SSS Number", 4, True),
        ("TIN Number", 4, True),
        ("Date Hired", 4, True),
    ]
    
    for title, row, bold in headers:
        ws[f'A{row}'] = title
        if bold:
            ws[f'A{row}'].font = Font(bold=True)

    employees = EmployeeLogin.objects.filter(is_active=True).exclude(is_admin=True).exclude(is_approved="Disapproved").exclude(is_approved="-")
    personal_infos = PersonalInformation.objects.filter(name__in=employees)
    emp_infos = EmploymentInformation.objects.filter(name__in=employees)

    for i, personal_info in enumerate(personal_infos, start=5):
        emp_info = emp_infos.filter(name=personal_info.name).first()

        contact_person = f'{personal_info.contact_firstname} {personal_info.contact_middlename} {personal_info.contact_lastname}'
        contact_address = f'{personal_info.contact_street}, {personal_info.contact_baranggay}, {personal_info.contact_city}, {personal_info.contact_province}'

        ws[f'A{i}'] = personal_info.name.idnumber if personal_info.name else ''
        ws[f'B{i}'] = personal_info.name.name if personal_info.name else ''
        ws[f'C{i}'] = personal_info.nickname if personal_info.nickname else ''
        ws[f'D{i}'] = emp_info.department.abreviation if emp_info and emp_info.department else ''
        ws[f'E{i}'] = emp_info.position.name if emp_info and emp_info.position else ''
        ws[f'F{i}'] = contact_person if contact_person else ''
        ws[f'G{i}'] = contact_address if contact_address else ''
        ws[f'H{i}'] = personal_info.contact_no if personal_info.contact_no else ''
        ws[f'I{i}'] = emp_info.sss_number if emp_info else ''
        ws[f'J{i}'] = emp_info.tin_number if emp_info else ''
        ws[f'K{i}'] = emp_info.date_hired.strftime('%m/%d/%Y') if emp_info and emp_info.date_hired else ''
  
    for row in ws['A4:K' + str(4 + len(employees) + 1)]:
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
     
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Employee_ID_Replacement.xlsx'
    wb.save(response)
    
    return response


@api_view(['GET'])
def department_percentages(request):
    total_employees = EmploymentInformation.objects.filter(name__is_active=True).count()

    if total_employees == 0:
        return Response({
            'labels': [],
            'percentages': [],
            'total_employees': total_employees
        })

    department_percentages = EmploymentInformation.objects.filter(
        name__is_active=True
    ).values(
        'department__abreviation'
    ).annotate(
        count=Count('id'),
        percentage=Cast(Count('id') * 100.0 / total_employees, FloatField())
    ).order_by('department__abreviation')

    labels = [dept['department__abreviation'] for dept in department_percentages if dept['department__abreviation']]
    percentages = [round(dept['percentage'], 1) for dept in department_percentages]

    return Response({
        'labels': labels,
        'percentages': percentages,
        'total_employees': total_employees
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def profile_completion_percentage(request):
    try:
        personal_info = PersonalInformation.objects.get(name=request.user)
        
        # List of all fields to check
        fields_to_check = [
            'nickname', 'gender', 'birth_date', 'birth_place',
            'contact_number', 'mother', 'father', 'work_email',
            'present_street', 'present_baranggay', 'present_city', 'present_province',
            'provincial_street', 'provincial_baranggay', 'provincial_city', 'provincial_province',
            'contact_firstname', 'contact_lastname',
            'contact_relation', 'contact_no',
            'contact_street', 'contact_baranggay', 'contact_city', 'contact_province'
        ]
        
        # Count how many fields are filled
        filled_fields = 0
        total_fields = len(fields_to_check)
        
        for field in fields_to_check:
            value = getattr(personal_info, field)
            if value is not None and value != '':
                filled_fields += 1
        
        # Calculate percentage
        percentage = int((filled_fields / total_fields) * 100)
        
        return Response({
            'percentage': percentage,
            'filled_fields': filled_fields,
            'total_fields': total_fields
        })
    
    except PersonalInformation.DoesNotExist:
        return Response({
            'percentage': 0,
            'filled_fields': 0,
            'total_fields': len(fields_to_check)
        })
    
@login_required(login_url='login')
def account_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect('accounts')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception as e:
            messages.error(request, f"Error loading file: {e}")
            return redirect('accounts')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                username = row[0]
                idnumber = row[1]
                firstname = row[2]
                lastname = row[3]
                name = f'{firstname} {lastname}'
                email = row[4]
                password = row[5]
                approver = row[6]

                # Validate required fields
                if not username or not idnumber or not email or not password:
                    messages.warning(request, f"Missing data in row: {row}. Skipping.")
                    continue

                # Validate email format
                try:
                    validate_email(email)
                except ValidationError:
                    messages.warning(request, f"Invalid email {email} in row: {row}. Skipping.")
                    continue

                # Check if ID number already exists
                if EmployeeLogin.objects.filter(idnumber=idnumber).exists():
                    messages.warning(request, f"Employee with ID {idnumber} already has an account. Skipping.")
                    continue 

                # Create the employee record
                employee = EmployeeLogin.objects.create(
                    username=username,
                    idnumber=idnumber,
                    firstname=firstname,
                    lastname=lastname,
                    name=name,
                    email=email,
                    password=make_password(password),
                    is_approver=approver
                )
            except Exception as e:
                messages.error(request, f"Error processing row {row}: {e}")

        messages.success(request, "Employee accounts imported successfully.")
        return redirect('accounts')

    return redirect('accounts')