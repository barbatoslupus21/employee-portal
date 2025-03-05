from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from UserLogin.models import EmployeeLogin
from .models import PerformanceEvaluation, PerformanceRouting, Assessment, Tasklist, EmployeeResponse, RecommendationsConcerns, TrainingRequest
from Profile.models import EmploymentInformation
from Settings.models import Status
from Notification.models import Notification
from django.utils import timezone
from datetime import date, datetime
from django.dispatch import receiver
from django.db.models.signals import post_save
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from openpyxl import Workbook
from django.db import IntegrityError

@receiver(post_save, sender=PerformanceEvaluation)
def create_evaluations_for_users(sender, instance, created, **kwargs):
    if created:
        regular_status = Status.objects.get(name="Regular")
        employees = EmployeeLogin.objects.filter(
            is_admin=False,
            is_active=True,
            employment_info__Status=regular_status
        )
        
        for employee in employees:
            EmployeeResponse.objects.create(employee=employee, quarter=instance)
            Notification.objects.create(
                level="Medium",
                module='Performance Evaluation',
                notifier=instance.created_by,
                page="performance-evaluation",
                reciever=employee,
                message='has posted the quarterly performance evaluation.'
            )

@login_required(login_url='login')
def admin_evaluation(request):
    if not request.user.is_hr:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    today = date.today()
    start_date = timezone.make_aware(datetime(today.year, 5, 1))
    end_date = timezone.make_aware(datetime(today.year + 1, 4, 30))

    quarters = PerformanceEvaluation.objects.order_by('-submitted_at')
    context = {
        'quarters': quarters
    }
    return render(request, 'Evaluation/admin-evaluation.html', context)
    
@login_required(login_url='login')
def create_evaluation(request):
    if not request.user.is_hr:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    if request.method == "POST":
        quarter = request.POST.get('quarter')
        period = request.POST.get('period')

        PerformanceEvaluation.objects.create(
            quarter=quarter,
            period=period,
            created_by=request.user
        )

        messages.success(request, 'Performance evaluation created.')
        return redirect('admin-performance')
    return redirect('admin-performance')

@login_required(login_url='login')
def delete_evaluation(request, pk):
    if not request.user.is_hr:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    selectedEval = get_object_or_404(PerformanceEvaluation, id=pk)
    if request.method == "POST":
        selectedEval.delete()
        messages.success(request, 'Performance evaluation deleted.')
        return redirect('admin-performance')
    return redirect('admin-performance')

@login_required(login_url='login')
def edit_evaluation(request, pk):
    if not request.user.is_hr:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    selectedEval = get_object_or_404(PerformanceEvaluation, id=pk)
    if request.method == "POST":
        quarter = request.POST.get('quarter')
        period = request.POST.get('period')
        selectedEval.quarter=quarter
        selectedEval.period=period
        selectedEval.save()
        messages.success(request, 'Saved Changes.')
        return redirect('admin-performance')
    return redirect('admin-performance')


@login_required(login_url='login')
def evaluation_view_response(request, pk):
    if not request.user.is_hr:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    employee_responses = EmployeeResponse.objects.filter(quarter__id=pk)

    employee_counts = EmployeeResponse.objects.filter(quarter__id=pk).count()
    formatted_employeeCount = f"{employee_counts:04}"

    submits = EmployeeResponse.objects.filter(quarter__id=pk, is_submitted=True).count()
    formatted_submits = f"{submits:04}"

    if employee_counts > 0:
        response_percentage = (submits / employee_counts) * 100
    else:
        response_percentage = 0

    quarter = get_object_or_404(PerformanceEvaluation, id=pk)
    context={
        'employee_responses':employee_responses,
        'quarter':quarter,
        'employee_counts':formatted_employeeCount,
        'submits':formatted_submits,
        'response_percentage':round(response_percentage, 2)
    }
    return render(request, 'Evaluation/admin-view-responses.html', context)

@login_required(login_url='login')
def close_survey(request, pk):
    if not request.user.is_hr:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    selectedEval = get_object_or_404(PerformanceEvaluation, id=pk)
    if request.method == "POST":
        selectedEval.state="CLOSED"
        selectedEval.save()
        messages.success(request, 'Evaluation officially closed.')
        return redirect('admin-performance')
    return redirect('admin-performance')

@login_required(login_url='login')
def import_tasklists(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect('admin-performance')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception as e:
            messages.error(request, f"Error loading file: {e}")
            return redirect('admin-performance')

        notified_employees = set()  # To track employees who have already been notified

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                idnumber = row[0]
                tasklist = row[1]
                employee = EmployeeLogin.objects.get(idnumber=idnumber)

                Tasklist.objects.create(
                    employee=employee,
                    tasklist=tasklist
                )

                if employee.id not in notified_employees:
                    Notification.objects.create(
                        level="Low",
                        module="Performance Evaluation",
                        notifier=request.user,
                        page="performance-evaluation",
                        reciever=employee,
                        message="uploaded your tasklist."
                    )
                    notified_employees.add(employee.id) 

            except EmployeeLogin.DoesNotExist:
                messages.error(request, f"Employee with ID {idnumber} does not exist.")
            except IntegrityError:
                messages.error(request, f"Error: Tasklist entry for {idnumber} already exists.")
            except Exception as e:
                messages.error(request, f"Error processing employee {idnumber}: {e}")

        messages.success(request, "Tasklist uploaded successfully.")
        return redirect('admin-performance')

    return redirect('admin-performance')

@login_required(login_url='login')
def view_employee_response(request, pk):
    if not request.user.is_hr:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))

    evaluations = Assessment.objects.filter(response__id=pk)
    recommendation = RecommendationsConcerns.objects.filter(response__id=pk).first()
    training = TrainingRequest.objects.filter(response__id=pk).first()
    responder = EmployeeResponse.objects.get(id=pk)
    context={
        'evaluations':evaluations,
        'responder':responder,
        'recommendation':recommendation,
        'training':training
    }
    return render(request, 'Evaluation/employee-evaluation-view.html', context)

@login_required(login_url='login')
def evaluation_view(request):
    today = date.today()
    start_date = timezone.make_aware(datetime(today.year, 5, 1))
    end_date = timezone.make_aware(datetime(today.year + 1, 4, 30))

    quarters = EmployeeResponse.objects.filter(employee=request.user).order_by('-posted_at')

    context = {
        'quarters':quarters
    }
    return render(request, 'Evaluation/performance-evaluation.html', context)

@login_required(login_url='login')
def evaluation_form(request, pk):
    quarter = get_object_or_404(EmployeeResponse, id=pk, employee=request.user)
    tasklists = Tasklist.objects.filter(employee=request.user)
    if not tasklists.exists():
        messages.error(request, "You currently have no tasklists assigned.")
        return redirect('performance-evaluation')
    
    context={
        'tasklists':tasklists,
        'quarter':quarter
    }
    return render(request, 'Evaluation/evaluation-form.html', context)

@login_required(login_url='login')
def submit_evaluation_form(request, pk):
    quarter = get_object_or_404(EmployeeResponse, id=pk, employee=request.user)
    tasklists = Tasklist.objects.filter(employee=request.user)
    approver = EmploymentInformation.objects.get(name=request.user)
    if not tasklists.exists():
        messages.error(request, "You currently have no tasklists assigned.")
        return redirect('performance-evaluation')
    
    if request.method == 'POST':
        training = request.POST.get('training_name')
        objective = request.POST.get('training_objective')

        quarter.date_submitted=timezone.now()
        quarter.is_submitted=True
        quarter.save()

        for tasklist in tasklists:
            assessment_value = request.POST.get(f'Assessment{tasklist.id}')
            
            if assessment_value:
                Assessment.objects.create(
                    response=quarter,
                    employee=request.user,
                    tasklist=tasklist.tasklist,
                    self_assessment=assessment_value
                )
        PerformanceRouting.objects.create(
            evaluation=quarter,
            approver=approver.approver,
        )

        if training:
            TrainingRequest.objects.create(
                response=quarter,
                training=training,
                objective=objective
            )

        Notification.objects.create(
            level="Medium",
            module="Performance Evaluation",
            notifier=request.user,
            page="supervisor-evaluation",
            reciever=approver.approver,
            message=f'has submitted a self-assessment for {quarter.quarter.quarter.lower()}.'
        )
        messages.success(request, "Your self-assessment has been submitted successfully.")
        return redirect('performance-evaluation')

    context = {
        'tasklists': tasklists,
        'quarter': quarter
    }
    return render(request, 'Evaluation/evaluation-form.html', context)

@login_required(login_url='login')
def supervisor_evaluation_view(request):
    today = date.today()
    start_date = timezone.make_aware(datetime(today.year, 5, 1))
    end_date = timezone.make_aware(datetime(today.year + 1, 4, 30))

    quarters = PerformanceEvaluation.objects.order_by('-submitted_at')

    level = EmploymentInformation.objects.get(name=request.user)
    quarter_percentages = {}

    for quarter in quarters:
        responses = EmployeeResponse.objects.filter(quarter=quarter, employee__employment_info__approver=request.user)

        total_responses = responses.count()
        submitted_responses = responses.filter(is_submitted=True).count()

        if total_responses > 0:
            quarter.percentage = (submitted_responses / total_responses) * 100
        else:
            quarter.percentage = 0

    context = {
        'quarters': quarters,
        'level':level
    }
    return render(request, 'Evaluation/supervisor-view.html', context)

@login_required(login_url='login')
def supervisor_view_quarter(request, pk):
    responses = EmployeeResponse.objects.filter(quarter__id=pk, employee__employment_info__approver=request.user)
    quarter = PerformanceEvaluation.objects.get(id=pk)
    status = PerformanceRouting.objects.filter(evaluation__quarter__id=pk, approver=request.user)
    evaluations = PerformanceRouting.objects.filter(evaluation__quarter__pk=pk, approver=request.user)
    level = EmploymentInformation.objects.get(name=request.user)

    context={
        'responses':responses,
        'quarter':quarter,
        'status':status,
        'evaluations':evaluations,
        'level':level
    }
    return render(request, 'Evaluation/supervisor-view-quarter.html', context)

@login_required(login_url='login')
def supervisor_assessment(request, pk):
    employee = EmployeeResponse.objects.get(id=pk)
    tasklists = Assessment.objects.filter(response__id=pk)
    level = EmploymentInformation.objects.get(name=request.user)
    training = TrainingRequest.objects.filter(response=pk).first()
    recommendations = RecommendationsConcerns.objects.filter(response=pk).first()

    context={
        'tasklists':tasklists,
        'employee':employee,
        'level':level,
        'training':training,
        'recommendations':recommendations
    }
    return render(request, 'Evaluation/supervisor-assessment.html', context)

@login_required(login_url='login')
def supervisor_assessment_update(request, pk):
    quarter = get_object_or_404(EmployeeResponse, id=pk)
    selectedRouting = get_object_or_404(PerformanceRouting, evaluation=pk, approver=request.user, status="Pending")
    tasklists = Assessment.objects.filter(response=quarter)
    approver = EmploymentInformation.objects.get(name=request.user)
    selectedResponse = get_object_or_404(RecommendationsConcerns, response=quarter)

    if request.method == 'POST':
        strength = request.POST.get('strength')
        weakness = request.POST.get('weakness')
        training = request.POST.get('training')
        comment = request.POST.get('comment')
        emp_comment = request.POST.get('emp_comment')

        quarter.date_submitted = timezone.now()
        quarter.is_submitted = True
        quarter.save()

        selectedRouting.status = "Approved"
        selectedRouting.submitted_at = timezone.now()
        selectedRouting.updated_at = timezone.now()
        selectedRouting.save()

        for tasklist in tasklists:
            supervisor_assessment_value = request.POST.get(f'Assessment{tasklist.id}')
            
            if supervisor_assessment_value:
                tasklist.supervisor_assessment = supervisor_assessment_value
                tasklist.save()

        PerformanceRouting.objects.create(
            evaluation=quarter,
            approver=approver.approver,
        )

        RecommendationsConcerns.objects.create(
            response=quarter,
            strength=strength,
            weakness=weakness,
            training_required=training,
            comment=comment,
            emp_comment=emp_comment
        )

        Notification.objects.create(
            level="Low",
            module="Performance Evaluation",
            notifier=request.user,
            page="supervisor-evaluation",
            reciever=approver.approver,
            message=f'has submitted evaluation for {quarter.employee.name.lower()}.'
        )
            
        messages.success(request, "Supervisor assessment has been submitted successfully.")
        return redirect('supervisor-view-quarter', pk=quarter.quarter.id)
    
@login_required(login_url='login')
def manager_evaluation_view(request):
    today = date.today()
    start_date = timezone.make_aware(datetime(today.year, 5, 1))
    end_date = timezone.make_aware(datetime(today.year + 1, 4, 30))

    quarters = PerformanceEvaluation.objects.order_by('-submitted_at')
    context = {
        'quarters': quarters
    }
    return render(request, 'Evaluation/manager-view.html', context)

@login_required(login_url='login')
def manager_view_quarter(request, pk):
    responses = PerformanceRouting.objects.filter(evaluation__quarter__id=pk, approver=request.user)
    quarter = PerformanceEvaluation.objects.get(id=pk)

    context={
        'responses':responses,
        'quarter':quarter
    }
    return render(request, 'Evaluation/manager-view-quarter.html', context)

@login_required(login_url='login')
def manager_evaluation(request, pk):
    assessments = Assessment.objects.filter(response=pk)
    recommendation = RecommendationsConcerns.objects.filter(response=pk).first()
    training = TrainingRequest.objects.filter(response=pk).first()
    responder = EmployeeResponse.objects.filter(id=pk).first
    status = PerformanceRouting.objects.filter(evaluation=pk).first
    context={
        'assessments':assessments,
        'responder':responder,
        'recommendation':recommendation,
        'training':training,
        'status':status
    }
    return render(request, 'Evaluation/manager-evaluation.html', context)

@login_required(login_url='login')
def manager_evaluation_approval(request, pk):
    selectedEvaluation = get_object_or_404(PerformanceRouting, evaluation=pk, approver=request.user)
    selectedResponse = get_object_or_404(EmployeeResponse, id=pk)
    selectedRecommendation = get_object_or_404(RecommendationsConcerns, response=selectedResponse)
    
    if request.method == 'POST':
        action = request.POST.get('action') 
        overall_comment = request.POST.get('overall_comment') 

        if action == 'approve':
            selectedEvaluation.status = "Approved"
            message_text = "Performance evaluation approved"
        elif action == 'disapprove':
            selectedEvaluation.status = "Disapproved"
            message_text = "Performance evaluation disapproved"

        selectedEvaluation.submitted_at = timezone.now()
        selectedEvaluation.updated_at = timezone.now()
        selectedResponse.is_evaluated = True
        selectedRecommendation.manager_comment = overall_comment
        selectedResponse.save()
        selectedEvaluation.save()
        selectedRecommendation.save()

        Notification.objects.create(
                level="High",
                module="Performance Evaluation",
                notifier=request.user,
                page="performance-evaluation",
                reciever=selectedEvaluation.evaluation.employee,
                message=f'has {selectedEvaluation.status.lower()} your performance evaluation'
        )
        messages.success(request, message_text)
        return redirect('manager-view-quarter', pk=selectedEvaluation.evaluation.quarter.id)

    return redirect('manager-view-quarter', pk=selectedEvaluation.evaluation.quarter.id)

@login_required(login_url='login')
def export_data_to_excel(request, pk):
    quarter = get_object_or_404(PerformanceEvaluation, id=pk)
    workbook = Workbook()

    # Sheet 1: Average
    average_sheet = workbook.active
    average_sheet.title = "Average"
    
    # Add headers
    average_sheet['A1'] = "RYONAN ELECTRIC PHILIPPINES CORPORATION"
    average_sheet['A1'].font = Font(bold=True)
    average_sheet['A2'] = f"{quarter.quarter} Summary Report"
    average_sheet['A2'].font = Font(bold=True)
    
    # Add the column headers for the average sheet
    average_sheet.append(['Employee Name', 'Total Average'])

    # Get assessments for the selected quarter
    assessments = Assessment.objects.filter(response__quarter=pk, supervisor_assessment__isnull=False)
    
    employee_averages = {}
    
    for assessment in assessments:
        employee_name = assessment.employee.name
        supervisor_assessment = int(assessment.supervisor_assessment)
        
        # Initialize the employee's total and count if not already
        if employee_name not in employee_averages:
            employee_averages[employee_name] = {'total': 0, 'count': 0}
        
        employee_averages[employee_name]['total'] += supervisor_assessment
        employee_averages[employee_name]['count'] += 1

    # Calculate averages and populate the average sheet
    for employee_name, data in employee_averages.items():
        if data['count'] > 0:
            average_value = round(data['total'] / (data['count'] * 5) * 100, 2)  # Assuming scale is 1 to 5
            average_sheet.append([employee_name, average_value])
        else:
            average_sheet.append([employee_name, 'No Data'])

    # Sheet 2: Recommendations and Concerns
    recommendations_sheet = workbook.create_sheet(title="Recommendations and Concerns")
    
    # Add headers
    recommendations_sheet['A1'] = "RYONAN ELECTRIC PHILIPPINES CORPORATION"
    recommendations_sheet['A1'].font = Font(bold=True)
    recommendations_sheet['A2'] = f"{quarter.quarter} Summary Report"
    recommendations_sheet['A2'].font = Font(bold=True)

    # Get recommendations and concerns
    recommendations = RecommendationsConcerns.objects.filter(response__quarter=pk)
    
    # Add data to sheet
    recommendations_sheet.append(['Employee', 'Strength', 'Weakness', 'Training Required', 'Comment'])
    for recommendation in recommendations:
        recommendations_sheet.append([
            recommendation.response.employee.name,
            recommendation.strength,
            recommendation.weakness,
            recommendation.training_required,
            recommendation.comment,
        ])

    # Sheet 3: Training Request
    training_sheet = workbook.create_sheet(title="Training Request")
    
    # Add headers
    training_sheet['A1'] = "RYONAN ELECTRIC PHILIPPINES CORPORATION"
    training_sheet['A1'].font = Font(bold=True)
    training_sheet['A2'] = f"{quarter.quarter} Summary Report"
    training_sheet['A2'].font = Font(bold=True)

    # Get training requests
    training_requests = TrainingRequest.objects.filter(response__quarter=pk)
    
    # Add data to sheet
    training_sheet.append(['Employee', 'Training', 'Objective'])
    for training in training_requests:
        training_sheet.append([
            training.response.employee.name,
            training.training,
            training.objective,
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{quarter.quarter}_Summary_Report.xlsx"'
    workbook.save(response)
    return response