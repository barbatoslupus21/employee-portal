from django.shortcuts import render, redirect,get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from UserLogin.models import EmployeeLogin
from .models import Training, ParticipantResponse, TrainingForm, TrainingApproval
from Profile.models import EmploymentInformation
import json
from django.utils import timezone
from datetime import timedelta, datetime
from django.db.models.functions import TruncMonth
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import JsonResponse
from rest_framework.decorators import api_view
from django.db.models import Avg, F, Count, Q
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from django.http import HttpResponse
from Notification.models import Notification

@login_required(login_url='login')
def admin_training(request):
    if not request.user.is_hr:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    participants = EmployeeLogin.objects.all().exclude(is_admin=True).exclude(is_active=False)
    trainings = Training.objects.all()

    context = {
        'participants':participants,
        'trainings':trainings
    }
    return render(request, 'Training/admin-training.html', context)

@login_required(login_url='login')
def create_training(request):
    if request.method == 'POST':
        training = Training.objects.create(
            training_date=request.POST['training_date'],
            training_title=request.POST['training_title'],
            training_objective=request.POST['training_obj'],
            training_speaker=request.POST['training_speaker']
        )

        selected_participants = json.loads(request.POST['selected_participants'])

        for participant_id in selected_participants:
            employee = EmployeeLogin.objects.get(idnumber=participant_id)
            ParticipantResponse.objects.create(
                employee=employee,
                training=training,
                action=False 
            )
            Notification.objects.create(
                level="Medium",
                module="Training-effectiveness",
                notifier=request.user,
                page="training",
                reciever=employee,
                message="has posted the training survey for you to answer."
            )

        return redirect('admin-training')

    participants = EmployeeLogin.objects.all()
    return render(request, 'Training/admin-training.html', {'participants': participants})

@login_required(login_url='login')
def delete_training(request, pk):
    deleteTraining = get_object_or_404(Training, id=pk)

    if request.method == "POST":
        deleteTraining.delete()
        messages.success(request, 'Training deleted.')
        return redirect('admin-training')
    else:
        messages.error(request, 'Invalid request method.')
    return redirect('admin-training')

@login_required(login_url='login')
def admin_view_training(request, pk):
    selectedTraining = ParticipantResponse.objects.filter(training__id=pk)
    training = get_object_or_404(Training, id=pk)

    participants = ParticipantResponse.objects.filter(training=pk)
    for participant in participants:
        try:
            participant_department = EmploymentInformation.objects.filter(name=participant.employee).first()
            if participant_department:
                participant.department = participant_department.department.abreviation
            else:
                participant.department = None
            participant_line = EmploymentInformation.objects.filter(name=participant.employee).first()
            if participant_line:
                participant.line = participant_line.line.line
            else:
                participant.line = None
        except EmploymentInformation.DoesNotExist:
            participant.department = None
            participant.line = None

    training_responses = TrainingForm.objects.filter(evaluation_to__training__id=pk)

    if not training_responses.exists():
        messages.error(request, 'There are no survey responses available for this training.')
        return redirect('admin-training')
    
    total_participants = ParticipantResponse.objects.filter(training__id=pk).count()
    formatted_total_participants = f"{total_participants:03}"

    responded = ParticipantResponse.objects.filter(training__id=pk, action=True).count()
    formatted_responded =  f'{responded:03}'

    if total_participants > 0:
        response_percentage = (responded / total_participants) * 100
    else:
        response_percentage = 0

    context={
        'selectedTraining':selectedTraining,
        'training':training,
        'participants':participants,
        'total_participants':formatted_total_participants,
        'responded':formatted_responded,
        'response_percentage':round(response_percentage, 2)
    }

    return render(request, 'Training/admin-view-training.html', context)

@login_required(login_url='login')
def training(request):
    trainings = ParticipantResponse.objects.filter(employee=request.user)
    return render(request, 'Training/training.html', {'trainings':trainings})

@login_required(login_url='login')
def training_view(request, pk):
    selectedTraining = get_object_or_404(Training, id=pk)
    return render(request, 'Training/take-training-evaluation.html', {'selectedTraining':selectedTraining})

@login_required(login_url='login')
def participant_response(request, pk):
    response = get_object_or_404(ParticipantResponse, training=pk, employee=request.user)

    if request.method == 'POST':
        job_related = request.POST.get('job_related')
        explain_clearly = request.POST.get('explain_clearly')
        suitable_topic = request.POST.get('suitable_topic')

        clear_goals = request.POST.get('clear_goals')
        met_goals = request.POST.get('met_goals')
        easy_follow = request.POST.get('easy_follow')
        easy_understand = request.POST.get('easy_understand')

        speaker_knowledge = request.POST.get('speaker_knowledge')
        clear_communication = request.POST.get('communication')
        answered_questions = request.POST.get('questions')

        training_org = request.POST.get('training_org')
        facilities = request.POST.get('facilities')
        materials = request.POST.get('material')

        interest = request.POST.get('topic_interest')
        future_recommendations = request.POST.get('recommendation')
        related_subjects = request.POST.get('related_topic')

        app_work1 = request.POST.get('app_work1')
        app_work2 = request.POST.get('app_work2')
        target_date = request.POST.get('targetDate')
        actual_date = request.POST.get('actualDate')
        app_self1 = request.POST.get('app_life1')
        app_self2 = request.POST.get('app_life2')

        ParticipantsTrainingForm = TrainingForm(
            employee=request.user,
            evaluation_to=response,
            job_related=job_related,
            explain_clearly=explain_clearly,
            suitable_topic=suitable_topic,
            clear_goals=clear_goals,
            met_goals=met_goals,
            easy_follow=easy_follow,
            easy_understand=easy_understand,
            speaker_knowledge=speaker_knowledge,
            clear_communication=clear_communication,
            answered_questions=answered_questions,
            training_org=training_org,
            facilities=facilities,
            materials=materials,
            interest=interest,
            future_recommendations=future_recommendations,
            related_subjects=related_subjects,
            app_work1=app_work1,
            app_work2=app_work2,
            target_date=target_date,
            actual_date=actual_date,
            app_self1=app_self1,
            app_self2=app_self2
        )

        try:
            employment_info = request.user.employment_info.get()
            
            Notification.objects.create(
                level='Medium',
                module='Training',
                notifier=request.user,
                page="supervisor-view",
                reciever=employment_info.approver,
                message='has submitted their training survey for evaluation.'
            )
        except EmploymentInformation.DoesNotExist:
            pass

        response.action=True
        response.save()
        ParticipantsTrainingForm.save()
        messages.success(request, 'Thank you for completing the survey.')
        return redirect('training')
    
    return render(request, 'Training/training.html')

@login_required(login_url='login')
def survey_response_view(request, pk):
    try:
        selectedSurvey = TrainingForm.objects.get(evaluation_to=pk)
    except TrainingForm.DoesNotExist:
        messages.error(request, "Please complete the survey for this training.")
        return redirect('training')
    return render(request, 'Training/training-view.html', {'selectedSurvey': selectedSurvey})

@login_required(login_url='login')
def close_training(request, pk):
    selectedTraining = get_object_or_404(Training, id=pk)

    if request.method == 'POST':
        selectedTraining.is_closed = True
        selectedTraining.save()
        messages.success(request, 'Training closed.')
        return redirect('admin-training')
    return redirect('admin-training')

@api_view(['GET'])
def training_content(request, pk):
    fields = ['job_related', 'explain_clearly', 'suitable_topic']
    
    averages = TrainingForm.objects.filter(evaluation_to__training__id=pk).aggregate(
        **{f'{field}_avg': Avg(field) for field in fields}
    )

    percentages = {field: round((averages[f'{field}_avg'] / 5) * 100, 2) if averages[f'{field}_avg'] is not None else 0 
                   for field in fields}    
    return JsonResponse(percentages)

@api_view(['GET'])
def training_structure(request, pk):
    fields = ['clear_goals', 'met_goals', 'easy_follow', 'easy_understand']
    
    averages = TrainingForm.objects.filter(evaluation_to__training__id=pk).aggregate(
        **{f'{field}_avg': Avg(field) for field in fields}
    )

    percentages = {field: round((averages[f'{field}_avg'] / 5) * 100, 2) if averages[f'{field}_avg'] is not None else 0 
                   for field in fields}   
    return JsonResponse(percentages)

@api_view(['GET'])
def training_speaker(request, pk):
    fields = ['speaker_knowledge', 'clear_communication', 'answered_questions']
    
    averages = TrainingForm.objects.filter(evaluation_to__training__id=pk).aggregate(
        **{f'{field}_avg': Avg(field) for field in fields}
    )

    percentages = {field: round((averages[f'{field}_avg'] / 5) * 100, 2) if averages[f'{field}_avg'] is not None else 0 
                   for field in fields}   
    return JsonResponse(percentages)

@api_view(['GET'])
def training_resources(request, pk):
    fields = ['training_org', 'facilities', 'materials']
    
    averages = TrainingForm.objects.filter(evaluation_to__training__id=pk).aggregate(
        **{f'{field}_avg': Avg(field) for field in fields}
    )

    percentages = {field: round((averages[f'{field}_avg'] / 5) * 100, 2) if averages[f'{field}_avg'] is not None else 0 
                   for field in fields}   
    return JsonResponse(percentages)

@login_required(login_url='login')
def delete_participant(request, pk):
    deleteParticipant = get_object_or_404(ParticipantResponse, id=pk)

    if request.method == "POST":
        deleteParticipant.delete()
        messages.success(request, 'Participant deleted.')
        return redirect('admin-training')
    else:
        messages.error(request, 'Invalid request method.')
    return redirect('admin-training')

@login_required(login_url='login')
def admin_response_view(request, pk):
    selectedSurvey = get_object_or_404(TrainingForm, employee=pk)
    return render(request, 'Training/training-view.html', {'selectedSurvey':selectedSurvey})

@login_required(login_url='login')
def export_training_excel(request, pk):
    selectedQuarter = ParticipantResponse.objects.filter(training__id=pk)
    training = get_object_or_404(Training, id=pk)
    wb = Workbook()
    
    # Summary Sheet
    ws = wb.active
    ws.title = "Summary"
    ws['A1'] = "Ryonan Electric Philippines Corporation"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = f"Training Survey Report for {training.training_title}"
    ws['A4'] = "Fields"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = "Average Percentage"
    ws['B4'].font = Font(bold=True)
    
    training_fields = [f.name for f in TrainingForm._meta.get_fields() if f.name not in ['id', 'employee', 'evaluation_to', 'interest', 'future_recommendations', 'related_subjects', 'app_work1', 'app_work2', 'target_date', 'actual_date', 'app_self1', 'app_self2', 'result_impact', 'recommendation', 'assessment', 'updated_at', 'submitted_at']]
    average_scores = TrainingForm.objects.filter(evaluation_to__training__id=pk).aggregate(
        **{f'{field}_avg': Avg(field) for field in training_fields}
    )
    
    for i, field in enumerate(training_fields, start=5):
        ws[f'A{i}'] = field.replace('_', ' ').title()
        ws[f'B{i}'] = f"{round((average_scores[f'{field}_avg'] / 5) * 100, 2)}%"
    
    for row in ws['A4:B' + str(4 + len(training_fields))]:
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Participants Sheet
    ws = wb.create_sheet("Participants")
    ws['A1'] = "Ryonan Electric Philippines Corporation"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = f"Training Survey Report for {training.training_title}"
    ws['A3'] = "Participants"
    ws['A5'] = "ID number"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = "Name"
    ws['B5'].font = Font(bold=True)
    ws['C5'] = "Department"
    ws['C5'].font = Font(bold=True)
    ws['D5'] = "Line"
    ws['D5'].font = Font(bold=True)
    ws['E5'] = "Email"
    ws['E5'].font = Font(bold=True)
    
    respondents = ParticipantResponse.objects.filter(training__id=pk, action=True)
    for respondent in respondents:
        try:
            respondent_department = EmploymentInformation.objects.filter(name=respondent.employee).first()
            if respondent_department:
                respondent.department = respondent_department.department.abreviation
            else:
                respondent.department = None
            respondent_line = EmploymentInformation.objects.filter(name=respondent.employee).first()
            if respondent_line:
                respondent.line = respondent_department.line.line
            else:
                respondent.line = None
        except EmploymentInformation.DoesNotExist:
            respondent.department = None
            respondent.line = None

    for i, respondent in enumerate(respondents, start=6):
        ws[f'A{i}'] = respondent.employee.idnumber
        ws[f'B{i}'] = respondent.employee.name
        ws[f'C{i}'] = respondent.department
        ws[f'D{i}'] = respondent.line
        ws[f'E{i}'] = respondent.employee.email
    
    for row in ws['A5:E' + str(5 + len(respondents))]:
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Survey Sheet
    ws = wb.create_sheet("Survey Results")
    ws['A1'] = "Ryonan Electric Philippines Corporation"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = f" Survey Results for {training.training_title}"
    ws['A4'] = "ID number"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = "Name"
    ws['B4'].font = Font(bold=True)
    ws['C4'] = "Department"
    ws['C4'].font = Font(bold=True)
    ws['D4'] = "Line"
    ws['D4'].font = Font(bold=True)
    ws['E4'] = "Email"
    ws['E4'].font = Font(bold=True)
    ws['F4'] = "Most Interesting"
    ws['F4'].font = Font(bold=True)
    ws['G4'] = "Feedback for Improvement"
    ws['G4'].font = Font(bold=True)
    ws['H4'] = "Future Learning Interests"
    ws['H4'].font = Font(bold=True)
    ws['I4'] = "New Learning to Enhance Job Effectiveness"
    ws['I4'].font = Font(bold=True)
    ws['J4'] = "Applying Training Insights to Improve Work Performance"
    ws['J4'].font = Font(bold=True)
    ws['K4'] = "Planned Implementation Date"
    ws['K4'].font = Font(bold=True)
    ws['L4'] = "Actual Implementation Date"
    ws['L4'].font = Font(bold=True)
    ws['M4'] = "New Skills for Use Beyond Work"
    ws['M4'].font = Font(bold=True)
    ws['N4'] = "Applying Training Insights to Daily Life"
    ws['N4'].font = Font(bold=True)
    ws['O4'] = "Supervisor's Result and Impact"
    ws['O4'].font = Font(bold=True)
    ws['P4'] = "Supervisor's Recommendation"
    ws['P4'].font = Font(bold=True)
    ws['Q4'] = "Overall Assessment"
    ws['Q4'].font = Font(bold=True)
    
    respondents = TrainingForm.objects.filter(evaluation_to__training__id=pk, evaluation_to__action=True)
    for respondent in respondents:
        try:
            respondent_department = EmploymentInformation.objects.filter(name=respondent.employee).first()
            if respondent_department:
                respondent.department = respondent_department.department.abreviation
            else:
                respondent.department = None
            respondent_line = EmploymentInformation.objects.filter(name=respondent.employee).first()
            if respondent_line:
                respondent.line = respondent_department.line.line
            else:
                respondent.line = None
        except EmploymentInformation.DoesNotExist:
            respondent.department = None
            respondent.line = None

    for i, respondent in enumerate(respondents, start=6):
        ws[f'A{i}'] = respondent.employee.idnumber
        ws[f'B{i}'] = respondent.employee.name
        ws[f'C{i}'] = respondent.department
        ws[f'D{i}'] = respondent.line
        ws[f'E{i}'] = respondent.employee.email
        ws[f'F{i}'] = respondent.interest
        ws[f'G{i}'] = respondent.future_recommendations
        ws[f'H{i}'] = respondent.related_subjects
        ws[f'I{i}'] = respondent.app_work1
        ws[f'J{i}'] = respondent.app_work2
        ws[f'K{i}'] = respondent.target_date
        ws[f'L{i}'] = respondent.actual_date
        ws[f'M{i}'] = respondent.app_self1
        ws[f'N{i}'] = respondent.app_self2
        ws[f'O{i}'] = respondent.result_impact
        ws[f'P{i}'] = respondent.recommendation
        ws[f'Q{i}'] = respondent.assessment
    
    for row in ws['A4:Q' + str(5 + len(respondents))]:
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
     
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Training_Report_{training.training_title}.xlsx'
    wb.save(response)
    
    return response

@login_required(login_url='login')
def supervisor_view(request):
    participants = ParticipantResponse.objects.filter(action=True, employee__employment_info__approver=request.user).order_by("-posted_at")
    current_date = timezone.now().date()
    for participant in participants:
        try:
            respondent_department = EmploymentInformation.objects.filter(name=participant.employee).first()
            if respondent_department:
                participant.department = respondent_department.department.abreviation
            else:
                participant.department = None

            respondent_line = EmploymentInformation.objects.filter(name=participant.employee).first()
            if respondent_line:
                participant.line = respondent_line.line.line
            else:
                participant.line = None

            training_date = participant.training.training_date
            six_months_from_training = training_date + timedelta(days=182)
            participant.is_six_months = (current_date == six_months_from_training)

        except EmploymentInformation.DoesNotExist:
            participant.department = None
            participant.line = None
    return render(request, 'Training/supervisor-training.html', {'participants':participants})

@login_required(login_url='login')
def evaluate_participant(request, pk):
    selectedParticipant = get_object_or_404(TrainingForm, evaluation_to=pk)
    department = EmploymentInformation.objects.get(name=selectedParticipant.employee)

    context={
        'selectedParticipant':selectedParticipant,
          'department':department
    }
    return render(request, 'Training/supervisor-evaluation.html', context)

@login_required(login_url='login')
def submit_evaluation(request, pk):
    selectedParticipant = get_object_or_404(TrainingForm, evaluation_to=pk)
    participant = get_object_or_404(ParticipantResponse, id=pk)
    approver = EmploymentInformation.objects.get(name=request.user)

    if request.method == "POST":
        result_impact = request.POST.get('supervisor_result')
        recommendation = request.POST.get('supervisor_recommendation')
        assessment = request.POST.get('assessment')

        selectedParticipant.result_impact = result_impact
        selectedParticipant.recommendation = recommendation
        selectedParticipant.assessment = assessment
        selectedParticipant.updated_at = timezone.now()
        participant.is_evaluated = True

        TrainingApproval.objects.create(
            training=participant,
            approver=approver.approver
        )

        selectedParticipant.save()
        participant.save()
        
        messages.success(request, 'Evaluation Submitted')
        return redirect('supervisor-view')

    return render(request, 'Training/supervisor-evaluation.html', {'selectedParticipant': selectedParticipant})

@login_required(login_url='login')
def manager_view(request):
    participants = TrainingApproval.objects.filter(training__action=True, approver=request.user)
    current_date = timezone.now().date()
    for participant in participants:
        try:
            respondent_department = EmploymentInformation.objects.filter(name=participant.training.employee).first()
            if respondent_department:
                participant.department = respondent_department.department.abreviation
            else:
                participant.department = None

            respondent_line = EmploymentInformation.objects.filter(name=participant.training.employee).first()
            if respondent_line:
                participant.line = respondent_line.line.line
            else:
                participant.line = None
            
        except EmploymentInformation.DoesNotExist:
            participant.department = None
            participant.line = None
    return render(request, 'Training/manager-training.html', {'participants':participants})

@login_required(login_url='login')
def manager_approval(request, pk):
    selectedSurvey = get_object_or_404(TrainingForm, evaluation_to=pk)
    return render(request, 'Training/manager-approval.html', {'selectedSurvey':selectedSurvey})

@login_required(login_url='login')
def manager_training_approval(request, pk):
    selectedTraining = get_object_or_404(ParticipantResponse, id=pk)
    if request.method == 'POST':
        action = request.POST.get('action') 

        if action == 'approve':
            selectedTraining.is_approved = "Approved"
            message_text = 'Training approved.'
        elif action == 'disapprove':
            selectedTraining.is_approved = "Disapproved"
            message_text = 'Training disapproved.'

        selectedTraining.save()
        messages.success(request, message_text)
        return redirect('manager-training-view')

    return redirect('manager-training-view')

@api_view(['GET'])
def training_counts(request):
    current_year = timezone.now().year
    current_month = timezone.now().month

    if current_month >= 5:
        start_date = datetime(current_year, 5, 1)
        end_date = datetime(current_year + 1, 4, 30)
    else:
        start_date = datetime(current_year - 1, 5, 1)
        end_date = datetime(current_year, 4, 30)

    monthly_counts = Training.objects.filter(
        training_date__range=[start_date, end_date]
    ).annotate(
        month=TruncMonth('training_date') 
    ).values('month').annotate(
        count=Count('id') 
    ).order_by('month')
    
    months = []
    counts = []
    
    for item in monthly_counts:
        months.append(item['month'].strftime('%b')) 
        counts.append(item['count'])
    
    return JsonResponse({
        'labels': months,
        'data': counts
    })