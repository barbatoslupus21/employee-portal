from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from UserLogin.models import EmployeeLogin
from .models import TicketCategory, TicketLevel, DeviceInformation, MISTickets
from django.db.models import Case, When, IntegerField
from django.utils import timezone
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from django.http import HttpResponse
from Profile.models import EmploymentInformation

@login_required(login_url='login')
def user_ticket(request):
    tickets = MISTickets.objects.filter(submitted_by=request.user).order_by("-date_submitted")
    userDevices = DeviceInformation.objects.filter(Assigned_to=request.user)
    levels = TicketLevel.objects.all()
    categories = TicketCategory.objects.all()

    context = { 
        'tickets':tickets,
        'userDevices':userDevices,
        'levels':levels,
        'categories': categories
    }
    return render(request, 'Ticketing/ticket.html' , context)

@login_required(login_url='login')
def submit_ticket(request):
    if request.method == "POST":
        device = request.POST.get('device')
        selectedDevice = DeviceInformation.objects.get(id=device)
        category = request.POST.get('category')
        selectedCategory = TicketCategory.objects.get(id=category)
        level = request.POST.get('level')
        selectedLevel = TicketLevel.objects.get(id=level)
        problem_details = request.POST.get('problem_details')

        MISTickets.objects.create(
            submitted_by=request.user,
            device=selectedDevice,
            category=selectedCategory,
            level=selectedLevel,
            problem_details=problem_details
        )

        messages.success(request, 'Your ticket successfully submitted')
        return redirect('ticket')
    return redirect('ticket')

@login_required(login_url='login')
def user_device(request):
    userDevices = DeviceInformation.objects.filter(Assigned_to=request.user)
    context = { 
        'userDevices':userDevices
    }
    return render(request, 'Ticketing/register-device.html' , context)

@login_required(login_url='login')
def edit_device(request, pk):
    selected = get_object_or_404(DeviceInformation, id=pk)

    if request.method == 'POST':
        Device = request.POST.get('Device')
        Device_code = request.POST.get('Device_code')
        Device_name = request.POST.get('Device_name')
        Device_location = request.POST.get('Device_location')

        selected.Device=Device
        selected.Device_Code=Device_code
        selected.Device_Name=Device_name
        selected.Device_location = Device_location
        selected.save()
        messages.success(request, 'Changes saved.')
        return redirect(request.META.get('HTTP_REFERER'))
    return redirect('user-devices')

@login_required(login_url='login')
def delete_device(request, pk):
    selected = get_object_or_404(DeviceInformation, id=pk)
    if request.method == 'POST':
        selected.delete()
        messages.success(request, 'Level deleted.')
        return redirect(request.META.get('HTTP_REFERER'))
    return redirect('user-devices')

@login_required(login_url='login')
def create_device(request):
    if request.method == 'POST':
        device = request.POST.get('newDevice')
        device_code = request.POST.get('newDevice_code')
        device_name = request.POST.get('newDevice_name')
        device_location = request.POST.get('newDevice_location')

        if DeviceInformation.objects.filter(Device_Code=device_code).exists():
            messages.error(request, 'Device already assigned.')
            return redirect(request.META.get('HTTP_REFERER'))

        DeviceInformation.objects.create(
            Assigned_to=request.user,
            Device=device,
            Device_Code=device_code,
            Device_Name=device_name,
            Device_location=device_location
        )
        messages.success(request, 'Device successfully registered.')
        return redirect(request.META.get('HTTP_REFERER'))
    
    return redirect('user-devices')

@login_required(login_url='login')
def admin_ticket(request):
    if not request.user.is_mis:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    tickets = MISTickets.objects.annotate(status_priority=Case(When(status="On Process", then=0),  default=1,output_field=IntegerField())).order_by('status_priority', '-date_submitted')

    total_tickets = MISTickets.objects.count()
    pending_tickets = MISTickets.objects.filter(status='On Process').count()
    
    first_ticket = MISTickets.objects.order_by('date_submitted').first()
    if first_ticket:
        first_ticket_date = first_ticket.date_submitted
        today = timezone.now().date()
        months_since_first_ticket = (today.year - first_ticket_date.year) * 12 + today.month - first_ticket_date.month
        if months_since_first_ticket > 0:
            avg_tickets_per_month = total_tickets / months_since_first_ticket
        else:
            avg_tickets_per_month = total_tickets 
    else:
        avg_tickets_per_month = 0

    avg_tickets_per_month_percentage = f"{avg_tickets_per_month:.1f}%"

    context={
        'tickets':tickets,
        'total_tickets':total_tickets,
        'pending_tickets':pending_tickets,
        'average_ticket':avg_tickets_per_month_percentage
    }
    return render(request, 'Ticketing/admin-ticket.html', context)

@login_required(login_url='login')
def submit_diagnosis(request, pk):
    selectedTicket = get_object_or_404(MISTickets, id=pk)
    if request.method == 'POST':
        status = request.POST.get('status')
        technician = request.POST.get('technician')
        diagnosis = request.POST.get('diagnosis')
        action = request.POST.get('action')
        reason = request.POST.get('reason')
        recommendation = request.POST.get('recommendation')

        selectedTicket.status=status
        selectedTicket.technician=technician
        selectedTicket.diagnosis=diagnosis
        selectedTicket.action_taken=action
        selectedTicket.possible_reason=reason
        selectedTicket.recommendation=recommendation
        selectedTicket.updated_at=timezone.now()
        selectedTicket.save()
        messages.success(request, 'Ticket diagnosed.')
        return redirect('admin-ticket')
    return redirect('admin-ticket')

@login_required(login_url='login')
def export_tickets(request):
    # Create a new workbook and active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "MIS Ticket"

    ws['A1'] = "Ryonan Electric Philippines Corporation"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = "Monthly IT Ticket Request Summary"
    headers = [
        "ID number", "Name", "Department", "Device", "Category", "Level",
        "Problem Detail", "Status", "Technician", "Diagnosis", 
        "Action Taken", "Possible Reason", "Recommendation", "Date Diagnosed"
    ]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}4'] = header
        ws[f'{col_letter}4'].font = Font(bold=True)

    current_date = timezone.now()
    tickets = MISTickets.objects.filter(
        date_submitted__month=current_date.month,
        date_submitted__year=current_date.year
    ).order_by('-date_submitted')

    emp_infos = EmploymentInformation.objects.filter(
        name__in=tickets.values_list('submitted_by', flat=True)
    ).select_related('department')
    emp_info_map = {info.name_id: info for info in emp_infos}

    for i, ticket in enumerate(tickets, start=5):
        emp_info = emp_info_map.get(ticket.submitted_by.id, None)

        ws[f'A{i}'] = ticket.submitted_by.idnumber or ''
        ws[f'B{i}'] = ticket.submitted_by.name or ''
        ws[f'C{i}'] = emp_info.department.abreviation if emp_info and emp_info.department else ''
        ws[f'D{i}'] = ticket.device.Device_Code if ticket.device else ''
        ws[f'E{i}'] = ticket.category.category if ticket.category else ''
        ws[f'F{i}'] = ticket.level.level if ticket.level else ''
        ws[f'G{i}'] = ticket.problem_details or ''
        ws[f'H{i}'] = ticket.status or ''
        ws[f'I{i}'] = ticket.technician or ''
        ws[f'J{i}'] = ticket.diagnosis or ''
        ws[f'K{i}'] = ticket.action_taken or ''
        ws[f'L{i}'] = ticket.possible_reason or ''
        ws[f'M{i}'] = ticket.recommendation or ''
        ws[f'N{i}'] = ticket.updated_at.strftime('%m/%d/%Y') if ticket.updated_at else ''

    for row in ws['A4:N' + str(4 + len(tickets))]:
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=MIS_ticket_request_{current_date.strftime("%B_%Y")}.xlsx'
    wb.save(response)

    return response


@login_required(login_url='login')
def admin_device(request):
    if not request.user.is_mis:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    Alldevices = DeviceInformation.objects.all()
    return render(request, 'Ticketing/admin-device.html', {'Alldevices':Alldevices})

@login_required(login_url='login')
def export_device_tickets(request, pk):
    device = get_object_or_404(DeviceInformation, id=pk)
    wb = Workbook()
    ws = wb.active

    # Set device information
    ws['A1'] = f"Device: {device.Device}"
    ws['A2'] = f"Device Code: {device.Device_Code}"
    ws['A3'] = f"Location: {device.Device_location}"
    ws['A4'] = f"Device User: {device.Assigned_to.name}"

    headers = ['Date Submitted', 'Category', 'Level', 'Problem Details']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True)

    tickets = MISTickets.objects.filter(device=device).order_by('-date_submitted')

    for row, ticket in enumerate(tickets, start=7):
        ws.cell(row=row, column=1, value=ticket.date_submitted.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row, column=2, value=ticket.category.category if ticket.category else '')
        ws.cell(row=row, column=3, value=ticket.level.level if ticket.level else '')
        ws.cell(row=row, column=4, value=ticket.problem_details)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Device_{device.Device_Code}_Tickets.xlsx'

    wb.save(response)

    return response