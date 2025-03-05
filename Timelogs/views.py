from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import EmployeeTimelogs
from UserLogin.models import EmployeeLogin
from .forms import ExcelUploadForm
from django.views import View
from openpyxl import load_workbook
from datetime import datetime, time
from .serializers import EmployeeTimelogsSerializer
from rest_framework import generics
from rest_framework.response import Response
from Profile.models import EmploymentInformation
import openpyxl

@login_required(login_url='login')
def admin_timelogs(request):
    if request.user.is_hr == False:
        messages.error(request, "You don't have the authority to open this section.")
        return redirect(request.META.get('HTTP_REFERER'))
    
    timelogs = EmployeeTimelogs.objects.all().order_by('-uploaded_at')

    for timelog in timelogs:
        try:
            employee = EmployeeLogin.objects.get(idnumber=timelog.employee_id)
            timelog.employee_name = employee.name
            department = EmploymentInformation.objects.get(name__idnumber=timelog.employee_id)
            timelog.department = department.department.abreviation
        except EmployeeLogin.DoesNotExist:
            timelog.employee_name = "Unknown" 

    return render(request, 'Timelogs/admin-timelogs.html', {'timelogs': timelogs})

@login_required(login_url='login')
def timelogs(request):
    return render(request, 'Timelogs/timelogs.html')

def format_time(time_value):
    if not time_value:
        return None
    
    if isinstance(time_value, time):
        return time_value.strftime("%I:%M %p")
    
    if isinstance(time_value, str):
        time_value = time_value.strip()
        try:
            time_obj = datetime.strptime(time_value, "%I:%M %p")
            return time_obj.strftime("%I:%M %p")
        except ValueError:
            try:
                time_obj = datetime.strptime(time_value, "%H:%M")
                return time_obj.strftime("%I:%M %p")
            except ValueError:
                return time_value
    return None

@login_required(login_url='login')
def upload_timelogs(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('file')

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect('admin-timelogs')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception as e:
            messages.error(request, f"Error loading file: {e}")
            return redirect('admin-timelogs')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                employee_id = row[0]
                log_date = row[1]
                time_in = row[2]
                time_out = row[3]

                # Convert log_date to date format
                if isinstance(log_date, str):
                    log_date = datetime.strptime(log_date, "%Y-%m-%d").date()
                elif isinstance(log_date, datetime):
                    log_date = log_date.date()

                # Convert time_in and time_out to time format
                if isinstance(time_in, str):
                    time_in = datetime.strptime(time_in, "%I:%M %p").time()  # Parses 12-hour time with AM/PM
                elif isinstance(time_in, datetime):
                    time_in = time_in.time()

                if isinstance(time_out, str):
                    time_out = datetime.strptime(time_out, "%I:%M %p").time()  # Parses 12-hour time with AM/PM
                elif isinstance(time_out, datetime):
                    time_out = time_out.time()

                # Retrieve the employee instance
                employee = EmployeeLogin.objects.get(idnumber=employee_id)

                # Create the timelog entry
                EmployeeTimelogs.objects.create(
                    employee_id=employee,
                    log_date=log_date,
                    time_in=time_in,
                    time_out=time_out
                )

            except EmployeeLogin.DoesNotExist:
                messages.error(request, f"Employee with ID {employee_id} does not exist.")
            except ValueError as ve:
                messages.error(request, f"Date or time format error for employee {employee_id}: {ve}")
            except Exception as e:
                messages.error(request, f"Error processing employee {employee_id}: {e}")

        messages.success(request, "Timelogs uploaded successfully.")
        return redirect('admin-timelogs')
    return redirect('admin-timelogs')


class EmployeeTimelogsListView(generics.ListAPIView):
    serializer_class = EmployeeTimelogsSerializer

    def get_queryset(self):
        user_idnumber = self.request.user
        return EmployeeTimelogs.objects.filter(employee_id=user_idnumber)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        timelogs = {}

        for log in queryset:
            log_date_str = log.log_date.strftime('%Y-%m-%d')
            
            if log_date_str not in timelogs:
                timelogs[log_date_str] = []

            timelogs[log_date_str].append({
                'timeIn': log.time_in,
                'timeOut': log.time_out,
            })

        return Response(timelogs) 
